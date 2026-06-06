#!/usr/bin/env python3
"""
REPORTE DE ASISTENCIA 1-TO-1
Sistema de reportes para empresas con clases individuales (Tekia, OHmyBox, etc.)
TuSpeaking - Enero 2026
"""

import sys
import mysql.connector
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de base de datos
DB_CONFIG = {
    'host': 'localhost',
    'database': 'aulatuspeaking35',
    'user': 'moodle35',
    'password': 'TuspeakingFix2025!'
}

# Estilos
HEADER_FILL = PatternFill('solid', fgColor='008BA3')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def conectar_db():
    return mysql.connector.connect(**DB_CONFIG)

def get_fill_porcentaje(porcentaje):
    if porcentaje >= 75:
        return GREEN_FILL
    elif porcentaje >= 50:
        return YELLOW_FILL
    return RED_FILL

def timestamp_to_date(ts):
    if ts and ts > 0:
        return datetime.fromtimestamp(ts).strftime('%d/%m/%Y')
    return '-'

def timestamp_to_datetime(ts):
    if ts and ts > 0:
        return datetime.fromtimestamp(ts).strftime('%d/%m/%Y %H:%M')
    return '-'

def formato_tiempo(minutos):
    """Convierte minutos a formato Xh Ym"""
    if minutos is None or minutos == 0:
        return '0m'
    minutos = int(minutos)
    horas = minutos // 60
    mins = minutos % 60
    if horas > 0 and mins > 0:
        return f'{horas}h {mins}m'
    elif horas > 0:
        return f'{horas}h'
    else:
        return f'{mins}m'

def obtener_alumnos_curso(cursor, courseid):
    cursor.execute("""
        SELECT DISTINCT
            u.id as userid,
            u.firstname,
            u.lastname,
            u.email,
            u.lastaccess
        FROM mdl_user u
        JOIN mdl_user_enrolments ue ON u.id = ue.userid
        JOIN mdl_enrol e ON ue.enrolid = e.id
        WHERE e.courseid = %s
        AND u.email NOT LIKE '%%tuspeaking%%'
        AND u.email NOT LIKE '%%fundae%%'
        AND u.suspended = 0
        ORDER BY u.lastname, u.firstname
    """, (courseid,))
    return cursor.fetchall()

def obtener_clases_alumno(cursor, studentid, email, acuity_type_pattern, fecha_inicio, fecha_fin):
    """Obtiene clases y tiempo de Zoom del alumno"""
    cursor.execute("""
        SELECT 
            COUNT(*) as total_clases,
            SUM(CASE WHEN zoom_clasecompletada = 1 THEN 1 ELSE 0 END) as clases_asistidas,
            COALESCE(SUM(CASE WHEN zoom_clasecompletada = 1 THEN zoom_duration ELSE 0 END), 0) as tiempo_zoom
        FROM mdl_i3code_acuityZoom
        WHERE studentid = %s
        AND acuity_type LIKE %s
        AND acuity_datetime >= %s
        AND acuity_datetime <= %s
    """, (studentid, acuity_type_pattern, fecha_inicio, fecha_fin + 'T23:59:59'))
    
    result = cursor.fetchone()
    total_clases = int(result[0] or 0)
    clases_asistidas = int(result[1] or 0)
    tiempo_zoom = int(result[2] or 0)
    
    if total_clases == 0:
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN zoom_clasecompletada = 1 THEN 1 ELSE 0 END) as asistidas,
                COALESCE(SUM(CASE WHEN zoom_clasecompletada = 1 THEN zoom_duration ELSE 0 END), 0) as tiempo
            FROM mdl_i3code_acuityZoom
            WHERE acuity_email = %s
            AND acuity_datetime >= %s
            AND acuity_datetime <= %s
        """, (email, fecha_inicio, fecha_fin + 'T23:59:59'))
        result = cursor.fetchone()
        if result:
            total_clases = int(result[0] or 0)
            clases_asistidas = int(result[1] or 0)
            tiempo_zoom = int(result[2] or 0)
    
    return total_clases, clases_asistidas, tiempo_zoom

def obtener_progreso_alumno(cursor, userid, courseid):
    """Calcula progreso usando actividades con seguimiento activado"""
    cursor.execute("""
        SELECT COUNT(*) 
        FROM mdl_course_modules cm
        WHERE cm.course = %s
        AND cm.completion > 0
        AND cm.deletioninprogress = 0
        AND cm.visible = 1
    """, (courseid,))
    total_modulos = int(cursor.fetchone()[0] or 0)
    
    if total_modulos == 0:
        return 0, 0, 0.0
    
    cursor.execute("""
        SELECT COUNT(*)
        FROM mdl_course_modules_completion cmc
        JOIN mdl_course_modules cm ON cmc.coursemoduleid = cm.id
        WHERE cm.course = %s
        AND cmc.userid = %s
        AND cmc.completionstate >= 1
        AND cm.completion > 0
        AND cm.deletioninprogress = 0
        AND cm.visible = 1
    """, (courseid, userid))
    completados = int(cursor.fetchone()[0] or 0)
    
    porcentaje = round((completados / total_modulos) * 100, 1) if total_modulos > 0 else 0.0
    
    return completados, total_modulos, porcentaje

def obtener_tiempo_plataforma(cursor, userid, courseid):
    """Calcula tiempo estimado en plataforma (acciones * 0.5 min)"""
    cursor.execute("""
        SELECT COUNT(*) as acciones
        FROM mdl_logstore_standard_log
        WHERE courseid = %s
        AND userid = %s
    """, (courseid, userid))
    
    result = cursor.fetchone()
    acciones = int(result[0] or 0) if result else 0
    tiempo_minutos = int(round(acciones * 0.5))
    
    return tiempo_minutos

def generar_reporte_empresa(cursor, empresa, fecha_inicio, fecha_fin, wb):
    ws = wb.create_sheet(title=empresa[:31])
    
    ws.merge_cells('A1:Q1')
    ws['A1'] = f'REPORTE DE ASISTENCIA 1-TO-1 - {empresa.upper()}'
    ws['A1'].font = Font(bold=True, size=14, color='008BA3')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:Q2')
    ws['A2'] = f'Período: {fecha_inicio} a {fecha_fin} | Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = [
        'Curso', 'Alumno', 'Email',
        'F.Inicio', 'F.Fin',
        'Clases C.', 'Clases R.', '% Asist.',
        'Total Act.', 'Complet.', '% Progreso',
        'T.Zoom', 'T.Plataf',
        '% Global', '% Total',
        'Última Entrada', 'FUNDAE'
    ]
    
    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER
    
    row = 5
    like_pattern = f'%{empresa}%'
    
    cursor.execute("""
        SELECT 
            oac.courseid,
            oac.classnmbr as clases_contratadas,
            oac.isfundae,
            c.fullname as curso,
            c.startdate,
            c.enddate,
            oat.acuitytype
        FROM own_acuity_course oac
        JOIN own_acuitytypes oat ON oac.acuityid = oat.acuityid
        JOIN mdl_course c ON oac.courseid = c.id
        WHERE oac.tipo_clase = '1TO1'
        AND (oat.acuitytype LIKE %s OR c.fullname LIKE %s)
        ORDER BY c.fullname
    """, (like_pattern, like_pattern))
    
    cursos = cursor.fetchall()
    
    totales = {
        'clases_contratadas': 0,
        'clases_realizadas': 0,
        'modulos_completados': 0,
        'modulos_totales': 0,
        'tiempo_zoom': 0,
        'tiempo_plataforma': 0,
        'alumnos': 0
    }
    
    for curso in cursos:
        courseid, clases_contratadas, isfundae, nombre_curso, startdate, enddate, acuity_type = curso
        clases_contratadas = int(clases_contratadas or 8)
        
        fecha_inicio_curso = timestamp_to_date(startdate)
        fecha_fin_curso = timestamp_to_date(enddate)
        
        alumnos = obtener_alumnos_curso(cursor, courseid)
        
        for alumno in alumnos:
            userid, firstname, lastname, email, lastaccess = alumno
            
            total_clases, clases_asistidas, tiempo_zoom = obtener_clases_alumno(
                cursor, userid, email, f'%{empresa}%', fecha_inicio, fecha_fin
            )
            
            pct_asistencia = round((clases_asistidas / clases_contratadas) * 100, 1) if clases_contratadas > 0 else 0.0
            
            mod_completados, mod_totales, pct_progreso = obtener_progreso_alumno(cursor, userid, courseid)
            
            tiempo_plataforma = obtener_tiempo_plataforma(cursor, userid, courseid)
            
            # Convertir a float para cálculos
            pct_asistencia = float(pct_asistencia)
            pct_progreso = float(pct_progreso)
            
            # % Global (50/50): 50% asistencia + 50% progreso plataforma
            pct_global = round((0.5 * pct_asistencia) + (0.5 * pct_progreso), 1)
            
            # % Total (80/20): 80% progreso plataforma + 20% asistencia clases
            pct_total = round((0.8 * pct_progreso) + (0.2 * pct_asistencia), 1)
            
            nombre_corto = nombre_curso[:32] + '...' if len(nombre_curso) > 32 else nombre_curso
            
            col = 1
            ws.cell(row=row, column=col, value=nombre_corto).border = BORDER
            col += 1
            ws.cell(row=row, column=col, value=f'{firstname} {lastname}').border = BORDER
            col += 1
            ws.cell(row=row, column=col, value=email).border = BORDER
            col += 1
            ws.cell(row=row, column=col, value=fecha_inicio_curso).border = BORDER
            col += 1
            ws.cell(row=row, column=col, value=fecha_fin_curso).border = BORDER
            col += 1
            ws.cell(row=row, column=col, value=clases_contratadas).border = BORDER
            col += 1
            ws.cell(row=row, column=col, value=clases_asistidas).border = BORDER
            col += 1
            
            cell = ws.cell(row=row, column=col, value=f'{pct_asistencia}%')
            cell.fill = get_fill_porcentaje(pct_asistencia)
            cell.border = BORDER
            col += 1
            
            ws.cell(row=row, column=col, value=mod_totales).border = BORDER
            col += 1
            ws.cell(row=row, column=col, value=mod_completados).border = BORDER
            col += 1
            
            cell = ws.cell(row=row, column=col, value=f'{pct_progreso}%')
            cell.fill = get_fill_porcentaje(pct_progreso)
            cell.border = BORDER
            col += 1
            
            # Tiempo Zoom en formato horas y minutos
            ws.cell(row=row, column=col, value=formato_tiempo(tiempo_zoom)).border = BORDER
            col += 1
            
            # Tiempo Plataforma en formato horas y minutos
            ws.cell(row=row, column=col, value=formato_tiempo(tiempo_plataforma)).border = BORDER
            col += 1
            
            cell = ws.cell(row=row, column=col, value=f'{pct_global}%')
            cell.fill = get_fill_porcentaje(pct_global)
            cell.border = BORDER
            col += 1
            
            cell = ws.cell(row=row, column=col, value=f'{pct_total}%')
            cell.fill = get_fill_porcentaje(pct_total)
            cell.border = BORDER
            col += 1
            
            ws.cell(row=row, column=col, value=timestamp_to_datetime(lastaccess)).border = BORDER
            col += 1
            
            fundae_valor = 'Sí' if isfundae == 't' else 'No'
            ws.cell(row=row, column=col, value=fundae_valor).border = BORDER
            
            totales['clases_contratadas'] += clases_contratadas
            totales['clases_realizadas'] += clases_asistidas
            totales['modulos_completados'] += mod_completados
            totales['modulos_totales'] += mod_totales
            totales['tiempo_zoom'] += tiempo_zoom
            totales['tiempo_plataforma'] += tiempo_plataforma
            totales['alumnos'] += 1
            
            row += 1
    
    # Fila de totales
    row += 1
    ws.cell(row=row, column=1, value='TOTALES').font = Font(bold=True)
    ws.cell(row=row, column=2, value=f'{totales["alumnos"]} alumnos').font = Font(bold=True)
    ws.cell(row=row, column=6, value=totales['clases_contratadas']).font = Font(bold=True)
    ws.cell(row=row, column=7, value=totales['clases_realizadas']).font = Font(bold=True)
    
    pct_total_asist = round((totales['clases_realizadas'] / totales['clases_contratadas']) * 100, 1) if totales['clases_contratadas'] > 0 else 0
    ws.cell(row=row, column=8, value=f'{pct_total_asist}%').font = Font(bold=True)
    
    ws.cell(row=row, column=9, value=totales['modulos_totales']).font = Font(bold=True)
    ws.cell(row=row, column=10, value=totales['modulos_completados']).font = Font(bold=True)
    
    pct_total_prog = round((totales['modulos_completados'] / totales['modulos_totales']) * 100, 1) if totales['modulos_totales'] > 0 else 0
    ws.cell(row=row, column=11, value=f'{pct_total_prog}%').font = Font(bold=True)
    
    ws.cell(row=row, column=12, value=formato_tiempo(totales['tiempo_zoom'])).font = Font(bold=True)
    ws.cell(row=row, column=13, value=formato_tiempo(totales['tiempo_plataforma'])).font = Font(bold=True)
    
    pct_global_total = round((0.5 * pct_total_asist) + (0.5 * pct_total_prog), 1)
    pct_total_total = round((0.8 * pct_total_prog) + (0.2 * pct_total_asist), 1)
    ws.cell(row=row, column=14, value=f'{pct_global_total}%').font = Font(bold=True)
    ws.cell(row=row, column=15, value=f'{pct_total_total}%').font = Font(bold=True)
    
    # Ajustar anchos
    anchos = [32, 20, 24, 10, 10, 9, 9, 9, 9, 9, 10, 10, 10, 9, 9, 15, 8]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    
    return totales

def generar_hoja_resumen(wb, empresas_totales):
    ws = wb.active
    ws.title = 'Resumen'
    
    ws.merge_cells('A1:J1')
    ws['A1'] = 'RESUMEN GLOBAL - REPORTES 1-TO-1'
    ws['A1'].font = Font(bold=True, size=14, color='008BA3')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    
    headers = ['Empresa', 'Alumnos', 'Clases C.', 'Clases R.', '% Asist.', '% Progreso', '% Global', '% Total', 'T.Zoom', 'T.Plataf']
    
    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    
    row = 5
    gran_total = {
        'alumnos': 0,
        'clases_contratadas': 0,
        'clases_realizadas': 0,
        'modulos_completados': 0,
        'modulos_totales': 0,
        'tiempo_zoom': 0,
        'tiempo_plataforma': 0
    }
    
    for empresa, totales in empresas_totales.items():
        ws.cell(row=row, column=1, value=empresa).border = BORDER
        ws.cell(row=row, column=2, value=totales['alumnos']).border = BORDER
        ws.cell(row=row, column=3, value=totales['clases_contratadas']).border = BORDER
        ws.cell(row=row, column=4, value=totales['clases_realizadas']).border = BORDER
        
        pct_asist = round((totales['clases_realizadas'] / totales['clases_contratadas']) * 100, 1) if totales['clases_contratadas'] > 0 else 0
        cell = ws.cell(row=row, column=5, value=f'{pct_asist}%')
        cell.fill = get_fill_porcentaje(pct_asist)
        cell.border = BORDER
        
        pct_prog = round((totales['modulos_completados'] / totales['modulos_totales']) * 100, 1) if totales['modulos_totales'] > 0 else 0
        cell = ws.cell(row=row, column=6, value=f'{pct_prog}%')
        cell.fill = get_fill_porcentaje(pct_prog)
        cell.border = BORDER
        
        pct_global = round((0.5 * pct_asist) + (0.5 * pct_prog), 1)
        cell = ws.cell(row=row, column=7, value=f'{pct_global}%')
        cell.fill = get_fill_porcentaje(pct_global)
        cell.border = BORDER
        
        pct_total = round((0.8 * pct_prog) + (0.2 * pct_asist), 1)
        cell = ws.cell(row=row, column=8, value=f'{pct_total}%')
        cell.fill = get_fill_porcentaje(pct_total)
        cell.border = BORDER
        
        ws.cell(row=row, column=9, value=formato_tiempo(totales['tiempo_zoom'])).border = BORDER
        ws.cell(row=row, column=10, value=formato_tiempo(totales['tiempo_plataforma'])).border = BORDER
        
        for key in gran_total:
            gran_total[key] += totales.get(key, 0)
        
        row += 1
    
    row += 1
    ws.cell(row=row, column=1, value='TOTAL GLOBAL').font = Font(bold=True)
    ws.cell(row=row, column=2, value=gran_total['alumnos']).font = Font(bold=True)
    ws.cell(row=row, column=3, value=gran_total['clases_contratadas']).font = Font(bold=True)
    ws.cell(row=row, column=4, value=gran_total['clases_realizadas']).font = Font(bold=True)
    
    pct_asist_t = round((gran_total['clases_realizadas'] / gran_total['clases_contratadas']) * 100, 1) if gran_total['clases_contratadas'] > 0 else 0
    ws.cell(row=row, column=5, value=f'{pct_asist_t}%').font = Font(bold=True)
    
    pct_prog_t = round((gran_total['modulos_completados'] / gran_total['modulos_totales']) * 100, 1) if gran_total['modulos_totales'] > 0 else 0
    ws.cell(row=row, column=6, value=f'{pct_prog_t}%').font = Font(bold=True)
    
    ws.cell(row=row, column=7, value=f'{round((0.5 * pct_asist_t) + (0.5 * pct_prog_t), 1)}%').font = Font(bold=True)
    ws.cell(row=row, column=8, value=f'{round((0.8 * pct_prog_t) + (0.2 * pct_asist_t), 1)}%').font = Font(bold=True)
    ws.cell(row=row, column=9, value=formato_tiempo(gran_total['tiempo_zoom'])).font = Font(bold=True)
    ws.cell(row=row, column=10, value=formato_tiempo(gran_total['tiempo_plataforma'])).font = Font(bold=True)
    
    # Leyenda
    row += 3
    ws.cell(row=row, column=1, value='LEYENDA:').font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value='% Global = 50% Asistencia + 50% Progreso')
    row += 1
    ws.cell(row=row, column=1, value='% Total = 80% Progreso + 20% Asistencia')
    row += 2
    ws.cell(row=row, column=1, value='≥75%')
    ws.cell(row=row, column=2).fill = GREEN_FILL
    ws.cell(row=row, column=2, value='Bueno')
    row += 1
    ws.cell(row=row, column=1, value='50-74%')
    ws.cell(row=row, column=2).fill = YELLOW_FILL
    ws.cell(row=row, column=2, value='Regular')
    row += 1
    ws.cell(row=row, column=1, value='<50%')
    ws.cell(row=row, column=2).fill = RED_FILL
    ws.cell(row=row, column=2, value='Bajo')
    
    anchos = [20, 10, 10, 10, 10, 12, 10, 10, 10, 10]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

def main():
    if len(sys.argv) < 3:
        print("Uso: python3 reporte_1to1.py <fecha_inicio> <fecha_fin> [empresa]")
        print("Ejemplo: python3 reporte_1to1.py 2025-09-01 2025-12-31")
        print("Ejemplo: python3 reporte_1to1.py 2025-09-01 2025-12-31 Tekia")
        sys.exit(1)
    
    fecha_inicio = sys.argv[1]
    fecha_fin = sys.argv[2]
    
    empresa_filtro = None
    for arg in sys.argv[3:]:
        if not arg.startswith('--'):
            empresa_filtro = arg
    
    print("=" * 60)
    print("  REPORTE DE ASISTENCIA 1-TO-1")
    print(f"  Período: {fecha_inicio} a {fecha_fin}")
    print(f"  Fecha generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    conn = conectar_db()
    cursor = conn.cursor()
    
    wb = Workbook()
    empresas_totales = {}
    
    empresas = [empresa_filtro] if empresa_filtro else ['Tekia', 'OHmyBox']
    
    for empresa in empresas:
        print(f"\nProcesando: {empresa}...")
        totales = generar_reporte_empresa(cursor, empresa, fecha_inicio, fecha_fin, wb)
        empresas_totales[empresa] = totales
        print(f"  - Alumnos: {totales['alumnos']}")
        print(f"  - Clases: {totales['clases_realizadas']}/{totales['clases_contratadas']}")
        if totales['modulos_totales'] > 0:
            pct_prog = round((totales['modulos_completados'] / totales['modulos_totales']) * 100, 1)
            pct_asist = round((totales['clases_realizadas'] / totales['clases_contratadas']) * 100, 1) if totales['clases_contratadas'] > 0 else 0
            print(f"  - % Asistencia: {pct_asist}%")
            print(f"  - % Progreso: {pct_prog}%")
            print(f"  - % Global (50/50): {round((0.5 * pct_asist) + (0.5 * pct_prog), 1)}%")
            print(f"  - % Total (80/20): {round((0.8 * pct_prog) + (0.2 * pct_asist), 1)}%")
            print(f"  - Tiempo Zoom: {formato_tiempo(totales['tiempo_zoom'])}")
            print(f"  - Tiempo Plataforma: {formato_tiempo(totales['tiempo_plataforma'])}")
    
    generar_hoja_resumen(wb, empresas_totales)
    
    if empresa_filtro:
        nombre_archivo = f'Reporte_1to1_{empresa_filtro}_{fecha_inicio}_a_{fecha_fin}.xlsx'
    else:
        nombre_archivo = f'Reporte_1to1_Empresas_{fecha_inicio}_a_{fecha_fin}.xlsx'
    
    ruta_archivo = f'/home/aulatuspeaking/www/app/moodle/reportes_1to1/{nombre_archivo}'
    
    wb.save(ruta_archivo)
    
    print(f"\n{'=' * 60}")
    print(f"Reporte generado: {ruta_archivo}")
    print(f"{'=' * 60}")
    
    cursor.close()
    conn.close()

if __name__ == "__main__":
    main()
