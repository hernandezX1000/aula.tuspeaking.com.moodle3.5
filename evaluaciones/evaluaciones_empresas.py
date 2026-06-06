#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IMPORTADOR DE CSV JOTFORM Y GENERADOR DE INFORMES
Sistema de Evaluaciones TuSpeaking
Fecha: 2025-12-29
"""

import sys
import os
import csv
import json
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import mysql.connector
    MYSQL_AVAILABLE = True
except ImportError:
    MYSQL_AVAILABLE = False

# CONFIGURACION
DB_CONFIG = {
    'host': 'localhost',
    'database': 'aulatuspeaking35',
    'user': 'moodle35',
    'password': 'TuspeakingFix2025!',
    'charset': 'utf8mb4'
}

OUTPUT_DIR = '/home/aulatuspeaking/www/app/moodle/reportes_evaluaciones'

ESCALA_NOTAS = {
    'poor(1)': 1, 'poor (1)': 1,
    'good(2)': 2, 'good (2)': 2,
    'good(3)': 3, 'good (3)': 3,
    'good(4)': 4, 'good (4)': 4,
    'excelent(5)': 5, 'excelent (5)': 5,
    'excellent(5)': 5, 'excellent (5)': 5,
}

ESCALA_FINAL = {
    'bad (less than 3)': 2, 'bad(less than 3)': 2,
    'not so good (3-4)': 3.5, 'not so good(3-4)': 3.5,
    'good (5-6)': 5.5, 'good(5-6)': 5.5, 'good ( 5- 6)': 5.5,
    'very good (7-8)': 7.5, 'very good(7-8)': 7.5, 'very good (7 - 8)': 7.5,
    'excelent (9-10)': 9.5, 'excelent(9-10)': 9.5, 'excelent (9 - 10)': 9.5,
    'excellent (9-10)': 9.5, 'excellent(9-10)': 9.5, 'excellent (9 - 10)': 9.5,
}

if OPENPYXL_AVAILABLE:
    GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF')
def get_db_connection():
    if not MYSQL_AVAILABLE:
        print("ERROR: mysql-connector no disponible")
        return None
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        return conn
    except mysql.connector.Error as e:
        print(f"Error de conexion: {e}")
        return None

def parse_puntuacion(texto):
    if not texto:
        return None
    texto_lower = texto.lower().strip()
    return ESCALA_NOTAS.get(texto_lower)

def parse_puntuacion_final(texto):
    if not texto:
        return None
    texto_lower = texto.lower().strip()
    return ESCALA_FINAL.get(texto_lower)

def calcular_nota_final(nota_grammar, nota_pron, nota_comm):
    notas = [n for n in [nota_grammar, nota_pron, nota_comm] if n is not None]
    if not notas:
        return None
    promedio = sum(notas) / len(notas)
    nota_final = (promedio - 1) * (10 - 1) / (5 - 1) + 1
    return round(nota_final, 2)

def determinar_certificado(nota_final):
    if nota_final is None:
        return 'ninguno'
    if nota_final >= 7.5:
        return 'superacion'
    elif nota_final >= 5:
        return 'participacion'
    return 'ninguno'

def get_fill_color(nota):
    if not OPENPYXL_AVAILABLE or nota is None:
        return None
    if nota >= 7.5:
        return GREEN_FILL
    elif nota >= 5:
        return YELLOW_FILL
    return RED_FILL

def parse_fecha(fecha_str):
    meses_es = {
        'ene': 'Jan', 'feb': 'Feb', 'mar': 'Mar', 'abr': 'Apr',
        'may': 'May', 'jun': 'Jun', 'jul': 'Jul', 'ago': 'Aug',
        'sep': 'Sep', 'oct': 'Oct', 'nov': 'Nov', 'dic': 'Dec'
    }
    fecha_normalizada = fecha_str.lower()
    for es, en in meses_es.items():
        fecha_normalizada = fecha_normalizada.replace(es, en.lower())
    
    formatos = ['%b %d, %Y', '%Y-%m-%d', '%d/%m/%Y']
    for formato in formatos:
        try:
            return datetime.strptime(fecha_normalizada.title(), formato)
        except ValueError:
            continue
    return datetime.now()
def importar_csv(archivo_csv, empresa_codigo='SERVIGUIDE'):
    print(f"\n{'='*60}")
    print(f"  IMPORTAR CSV DE EVALUACIONES")
    print(f"  Archivo: {archivo_csv}")
    print(f"  Empresa: {empresa_codigo}")
    print(f"{'='*60}\n")
    
    conn = get_db_connection()
    if not conn:
        return False
    
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT id FROM mdl_coding_empresas WHERE codigo = %s", (empresa_codigo,))
    empresa = cursor.fetchone()
    if not empresa:
        print(f"ERROR: Empresa {empresa_codigo} no encontrada")
        return False
    empresa_id = empresa['id']
    
    cursor.execute("""
        SELECT id FROM mdl_coding_ediciones 
        WHERE empresa_id = %s AND activo = 1 
        ORDER BY fecha_inicio DESC LIMIT 1
    """, (empresa_id,))
    edicion = cursor.fetchone()
    edicion_id = edicion['id'] if edicion else None
    
    registros_procesados = 0
    registros_error = 0
    
    with open(archivo_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        
        for row in reader:
            try:
                fecha_str = row.get('Submission Date', '')
                teacher = row.get('Teacher', '')
                student = row.get('Student', '')
                
                if not student:
                    continue
                
                print(f"  Procesando: {student}...")
                
                grammar_scores = []
                pronunciation_scores = []
                communicative_scores = []
                
                for key, value in row.items():
                    key_lower = key.lower()
                    score = parse_puntuacion(value)
                    
                    if 'grammar' in key_lower and 'vocabulary' in key_lower and score:
                        grammar_scores.append(score)
                    elif 'pronunciation' in key_lower and score:
                        pronunciation_scores.append(score)
                    elif 'communicative' in key_lower and score:
                        communicative_scores.append(score)
                
                nota_grammar = sum(grammar_scores) / len(grammar_scores) if grammar_scores else None
                nota_pron = sum(pronunciation_scores) / len(pronunciation_scores) if pronunciation_scores else None
                nota_comm = sum(communicative_scores) / len(communicative_scores) if communicative_scores else None
                
                oral_exam = None
                participation = None
                homework = None
                
                for key, value in row.items():
                    key_lower = key.lower()
                    if 'oral' in key_lower and 'exam' in key_lower:
                        oral_exam = parse_puntuacion_final(value)
                    elif 'participation' in key_lower:
                        participation = parse_puntuacion_final(value)
                    elif 'homework' in key_lower and 'status' in key_lower:
                        homework = parse_puntuacion_final(value)
                
                nota_final = calcular_nota_final(nota_grammar, nota_pron, nota_comm)
                
                should_move = row.get('Should the student', '')
                recommended_level = row.get('Recommended level:', '')
                
                recomendacion = 'mantener_nivel'
                if 'next level' in should_move.lower() or 'moved' in should_move.lower():
                    recomendacion = 'subir_nivel'
                
                justificacion = row.get('Justify your choices with insightful information. Please give details that show you know the student (+20 words)', '')
                homework_desc = row.get('Homework Description', '')
                spreadsheet = row.get('Spreadsheet', '')
                
                certificado_tipo = determinar_certificado(nota_final)
                submission_id = f"csv_{student}_{fecha_str}".replace(' ', '_')[:50]
                
                cursor.execute("""
                    SELECT id FROM mdl_coding_evaluaciones 
                    WHERE alumno_nombre = %s AND empresa_id = %s
                """, (student, empresa_id))
                existente = cursor.fetchone()
                
                if existente:
                    cursor.execute("""
                        UPDATE mdl_coding_evaluaciones SET
                            profesor = %s, nivel_evaluado = %s,
                            nota_grammar_vocabulary = %s, nota_pronunciation = %s, nota_communicative = %s,
                            nota_oral_exam = %s, nota_participacion = %s, nota_homework = %s,
                            nota_final = %s, recomendacion = %s, nivel_recomendado = %s,
                            justificacion = %s, comentarios_homework = %s, resumen_spreadsheet = %s,
                            certificado_tipo = %s, fecha_modificacion = NOW()
                        WHERE id = %s
                    """, (teacher, 'B1', nota_grammar, nota_pron, nota_comm,
                          oral_exam, participation, homework, nota_final, 
                          recomendacion, recommended_level, justificacion, 
                          homework_desc, spreadsheet, certificado_tipo, existente['id']))
                else:
                    cursor.execute("""
                        INSERT INTO mdl_coding_evaluaciones (
                            empresa_id, edicion_id, jotform_submission_id, fecha_evaluacion,
                            profesor, alumno_nombre, nivel_evaluado,
                            nota_grammar_vocabulary, nota_pronunciation, nota_communicative,
                            nota_oral_exam, nota_participacion, nota_homework,
                            nota_final, recomendacion, nivel_recomendado,
                            justificacion, comentarios_homework, resumen_spreadsheet,
                            certificado_tipo, fecha_creacion
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                    """, (empresa_id, edicion_id, submission_id, parse_fecha(fecha_str),
                          teacher, student, 'B1', nota_grammar, nota_pron, nota_comm,
                          oral_exam, participation, homework, nota_final,
                          recomendacion, recommended_level, justificacion,
                          homework_desc, spreadsheet, certificado_tipo))
                
                registros_procesados += 1
                print(f"    OK Nota: {nota_final:.2f} | Cert: {certificado_tipo}")
                
            except Exception as e:
                print(f"    ERROR: {e}")
                registros_error += 1
    
    conn.commit()
    cursor.close()
    conn.close()
    
    print(f"\n{'='*60}")
    print(f"  RESUMEN: Procesados={registros_procesados}, Errores={registros_error}")
    print(f"{'='*60}\n")
    return True
def generar_informe(empresa_codigo='SERVIGUIDE'):
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl no disponible")
        return None
    
    print(f"\n{'='*60}")
    print(f"  GENERAR INFORME DE EVALUACIONES")
    print(f"  Empresa: {empresa_codigo}")
    print(f"{'='*60}\n")
    
    conn = get_db_connection()
    if not conn:
        return None
    
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT alumno_nombre, profesor, nivel_evaluado,
               nota_grammar_vocabulary, nota_pronunciation, nota_communicative,
               nota_oral_exam, nota_participacion, nota_homework,
               nota_final, recomendacion, nivel_recomendado, justificacion, certificado_tipo
        FROM mdl_coding_evaluaciones e
        JOIN mdl_coding_empresas emp ON e.empresa_id = emp.id
        WHERE emp.codigo = %s
        ORDER BY alumno_nombre
    """, (empresa_codigo,))
    evaluaciones = cursor.fetchall()
    
    if not evaluaciones:
        print("No hay evaluaciones")
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Evaluaciones"
    
    headers = ['Alumno', 'Profesor', 'Nivel', 'Grammar', 'Pronunciation', 'Communicative',
               'Nota Final', 'Certificado', 'Recomendacion', 'Nivel Recomendado', 'Justificacion']    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    for row_num, ev in enumerate(evaluaciones, 2):
        ws.cell(row=row_num, column=1, value=ev['alumno_nombre'])
        ws.cell(row=row_num, column=2, value=ev['profesor'])
        ws.cell(row=row_num, column=3, value=ev['nivel_evaluado'])
        if ev['nota_grammar_vocabulary']:
            nota_g = (float(ev['nota_grammar_vocabulary']) - 1) * 9 / 4 + 1
            ws.cell(row=row_num, column=4, value=round(nota_g, 2))
        if ev['nota_pronunciation']:
            nota_p = (float(ev['nota_pronunciation']) - 1) * 9 / 4 + 1
            ws.cell(row=row_num, column=5, value=round(nota_p, 2))
        if ev['nota_communicative']:
            nota_c = (float(ev['nota_communicative']) - 1) * 9 / 4 + 1
            ws.cell(row=row_num, column=6, value=round(nota_c, 2))
        if ev['nota_final']:
            cell = ws.cell(row=row_num, column=7, value=float(ev['nota_final']))
            cell.fill = get_fill_color(float(ev['nota_final']))
        ws.cell(row=row_num, column=8, value=ev['certificado_tipo'])
        ws.cell(row=row_num, column=9, value=ev['recomendacion'])
        ws.cell(row=row_num, column=10, value=ev['nivel_recomendado'])
        ws.cell(row=row_num, column=11, value=ev['justificacion'])    
    for i, width in enumerate([30, 20, 10, 12, 12, 12, 12, 15, 15, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    fecha = datetime.now().strftime('%Y-%m-%d')
    archivo = f"{OUTPUT_DIR}/Informe_Evaluaciones_{empresa_codigo}_{fecha}.xlsx"
    wb.save(archivo)
    
    cursor.close()
    conn.close()
    
    print(f"Informe generado: {archivo}")
    print(f"Total evaluaciones: {len(evaluaciones)}")
    return archivo


def generar_certificados_csv(empresa_codigo='SERVIGUIDE', tipo='todos'):
    print(f"\n{'='*60}")
    print(f"  GENERAR CSV PARA CERTIFICADOS")
    print(f"  Empresa: {empresa_codigo}, Tipo: {tipo}")
    print(f"{'='*60}\n")
    
    conn = get_db_connection()
    if not conn:
        return None
    
    cursor = conn.cursor(dictionary=True)
    
    tipo_filter = ""
    if tipo == 'superacion':
        tipo_filter = "AND e.certificado_tipo = 'superacion'"
    elif tipo == 'participacion':
        tipo_filter = "AND e.certificado_tipo = 'participacion'"
    elif tipo == 'todos':
        tipo_filter = "AND e.certificado_tipo IN ('superacion', 'participacion')"
    
    cursor.execute(f"""
        SELECT e.alumno_nombre, e.nivel_evaluado, e.nivel_recomendado,
               e.nota_final, e.certificado_tipo,
               ed.fecha_inicio, ed.fecha_fin, ed.idioma, emp.nombre as empresa
        FROM mdl_coding_evaluaciones e
        JOIN mdl_coding_empresas emp ON e.empresa_id = emp.id
        LEFT JOIN mdl_coding_ediciones ed ON e.edicion_id = ed.id
        WHERE emp.codigo = %s {tipo_filter}
        ORDER BY e.certificado_tipo DESC, e.alumno_nombre
    """, (empresa_codigo,))
    evaluaciones = cursor.fetchall()
    
    if not evaluaciones:
        print("No hay evaluaciones para certificados")
        return None
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    fecha = datetime.now().strftime('%Y-%m-%d')
    archivo = f"{OUTPUT_DIR}/Certificados_{empresa_codigo}_{fecha}.csv"
    
    with open(archivo, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['nombre_completo', 'nivel', 'nivel_alcanzado', 'nota_final',
                        'fecha_inicio', 'fecha_fin', 'idioma', 'empresa', 'tipo_certificado'])
        for ev in evaluaciones:
            writer.writerow([
                ev['alumno_nombre'],
                ev['nivel_evaluado'],
                ev['nivel_recomendado'] or ev['nivel_evaluado'],
                f"{float(ev['nota_final']):.1f}" if ev['nota_final'] else '',
                ev['fecha_inicio'].strftime('%d/%m/%Y') if ev['fecha_inicio'] else '',
                ev['fecha_fin'].strftime('%d/%m/%Y') if ev['fecha_fin'] else '',
                ev['idioma'] or 'Ingles',
                ev['empresa'],
                'SUPERACION' if ev['certificado_tipo'] == 'superacion' else 'PARTICIPACION'
            ])
    
    cursor.close()
    conn.close()
    
    print(f"CSV generado: {archivo}")
    print(f"Total: {len(evaluaciones)} certificados")
    return archivo


def main():
    if len(sys.argv) < 2:
        print("""
USO:
    python3 evaluaciones_empresas.py importar archivo.csv [EMPRESA]
    python3 evaluaciones_empresas.py informe [EMPRESA]
    python3 evaluaciones_empresas.py certificados [EMPRESA] [tipo]

EJEMPLOS:
    python3 evaluaciones_empresas.py importar evaluaciones.csv SERVIGUIDE
    python3 evaluaciones_empresas.py informe SERVIGUIDE
    python3 evaluaciones_empresas.py certificados SERVIGUIDE todos
        """)
        return
    
    comando = sys.argv[1].lower()
    
    if comando == 'importar':
        archivo = sys.argv[2] if len(sys.argv) > 2 else None
        if not archivo:
            print("ERROR: Falta archivo CSV")
            return
        empresa = sys.argv[3] if len(sys.argv) > 3 else 'SERVIGUIDE'
        importar_csv(archivo, empresa)
    elif comando == 'informe':
        empresa = sys.argv[2] if len(sys.argv) > 2 else 'SERVIGUIDE'
        generar_informe(empresa)
    elif comando == 'certificados':
        empresa = sys.argv[2] if len(sys.argv) > 2 else 'SERVIGUIDE'
        tipo = sys.argv[3] if len(sys.argv) > 3 else 'todos'
        generar_certificados_csv(empresa, tipo)
    else:
        print(f"Comando desconocido: {comando}")

if __name__ == '__main__':
    main()
