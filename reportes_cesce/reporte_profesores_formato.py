#!/usr/bin/env python3
"""
REPORTE ASISTENCIA PROFESORES CESCE POR GRUPOS
Con detalle de registros Zoom mensualizados
"""

import sys
import mysql.connector
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from dateutil.relativedelta import relativedelta
from dateutil.parser import parse
from collections import defaultdict

fecha_inicio = sys.argv[1] if len(sys.argv) > 1 else '2025-09-01'
fecha_fin = sys.argv[2] if len(sys.argv) > 2 else '2025-12-19'
# === DETECCIÓN AUTOMÁTICA DE TABLAS ===
if fecha_inicio >= '2026-01-01':
    TABLA_CITAS = 'mdl_i3code_acuityZoom'
    TABLA_PART = 'mdl_i3code_acuityZoom_participants'
    CAMPO_NAME = 'zoom_name'
    CAMPO_EMAIL = 'zoom_email'
    CAMPO_DURATION = 'zoom_duration/60'
    print(f"  Usando tablas i3code (2026+)")
else:
    TABLA_CITAS = 'mdl_cesce_acuityZoom'
    TABLA_PART = 'mdl_coding_zoom_participants'
    CAMPO_NAME = 'participant_name'
    CAMPO_EMAIL = 'participant_email'
    CAMPO_DURATION = 'duration_minutes'
    print(f"  Usando tablas cesce (2025)")


print("=" * 60)
print("  REPORTE ASISTENCIA PROFESORES CESCE POR GRUPOS")
print(f"  Período: {fecha_inicio} a {fecha_fin}")
print(f"  Fecha generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 60)

conn = mysql.connector.connect(
    host='localhost',
    database='aulatuspeaking35',
    user='moodle35',
    password='TuspeakingFix2025!'
)
cursor = conn.cursor(dictionary=True)

# Obtener meses
inicio = parse(fecha_inicio)
fin = parse(fecha_fin)
meses = []
current = inicio.replace(day=1)
while current <= fin:
    meses.append(current.strftime('%Y-%m'))
    current += relativedelta(months=1)

meses_es = {
    '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
    '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
    '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dic'
}

meses_es_largo = {
    '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
    '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
    '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
}

def normalizar_profesor(nombre):
    if not nombre:
        return ''
    nombre = nombre.lower().strip()
    nombre = nombre.replace('(host)', '').replace('(co-host)', '').strip()
    return nombre

def es_mismo_profesor(asignado, presente):
    if not asignado or not presente:
        return False
    asignado = normalizar_profesor(asignado)
    presente = normalizar_profesor(presente)
    if asignado == presente:
        return True
    asignado_parts = asignado.split()
    presente_parts = presente.split()
    if asignado_parts and presente_parts:
        if asignado_parts[0] == presente_parts[0]:
            return True
    return False

# Obtener grupos únicos con profesor
sql_grupos = f"""
SELECT 
    acuity_type as grupo,
    acuity_calendar as profesor,
    COUNT(DISTINCT acuity_email) as num_alumnos
FROM {TABLA_CITAS}
WHERE acuity_datetime >= %s 
  AND acuity_datetime <= %s
  AND acuity_firstname != 'Booking'
GROUP BY acuity_type, acuity_calendar
ORDER BY acuity_calendar, acuity_type
"""
cursor.execute(sql_grupos, (fecha_inicio, fecha_fin))
grupos_info = {row['grupo']: {'profesor': row['profesor'], 'alumnos': row['num_alumnos']} for row in cursor.fetchall()}
print(f"Grupos encontrados: {len(grupos_info)}")

# Consulta: clases por grupo con TODOS los profesores tuspeaking presentes
sql_clases = f"""
SELECT 
    ca.acuity_type as grupo,
    ca.acuity_calendar as profesor_asignado,
    DATE(ca.acuity_datetime) as fecha,
    ca.zoom_meetingid,
    (SELECT GROUP_CONCAT(DISTINCT SUBSTRING_INDEX(p.{CAMPO_NAME}, ' (', 1) SEPARATOR ', ')
     FROM {TABLA_PART} p 
     WHERE p.zoom_meetingid = ca.zoom_meetingid 
     AND p.{CAMPO_EMAIL} LIKE '%%tuspeaking%%'
     AND p.{CAMPO_DURATION} >= 30
    ) as profesores_presentes,
    (SELECT COUNT(DISTINCT p.{CAMPO_EMAIL})
     FROM {TABLA_PART} p 
     WHERE p.zoom_meetingid = ca.zoom_meetingid 
     AND p.{CAMPO_EMAIL} LIKE '%%tuspeaking%%'
     AND p.{CAMPO_DURATION} >= 30
    ) as num_profesores
FROM {TABLA_CITAS} ca
WHERE ca.acuity_datetime >= %s 
  AND ca.acuity_datetime <= %s
  AND acuity_firstname != 'Booking'
GROUP BY ca.acuity_type, DATE(ca.acuity_datetime), ca.zoom_meetingid, ca.acuity_calendar
ORDER BY ca.acuity_type, ca.acuity_datetime
"""
cursor.execute(sql_clases, (fecha_inicio, fecha_fin))
clases_raw = cursor.fetchall()

# Consulta detallada para hojas de Zoom
sql_zoom_detalle = f"""
SELECT 
    ca.acuity_type as grupo,
    ca.acuity_calendar as profesor_asignado,
    DATE(ca.acuity_datetime) as fecha,
    TIME(ca.acuity_datetime) as hora,
    ca.zoom_meetingid,
    ca.acuity_duration as duracion_programada,
    p.{CAMPO_NAME},
    p.{CAMPO_EMAIL},
    p.{CAMPO_DURATION} as duracion_participante,
    CASE 
        WHEN p.{CAMPO_NAME} LIKE '%%(Host)%%' THEN 'Host'
        WHEN p.{CAMPO_NAME} LIKE '%%(Co-host)%%' THEN 'Co-host'
        WHEN p.{CAMPO_EMAIL} LIKE '%%tuspeaking%%' THEN 'Host'
        ELSE 'Participante'
    END as rol
FROM {TABLA_CITAS} ca
LEFT JOIN {TABLA_PART} p ON p.zoom_meetingid = ca.zoom_meetingid
WHERE ca.acuity_datetime >= %s 
  AND ca.acuity_datetime <= %s
  AND ca.acuity_firstname != 'Booking'
  AND (p.{CAMPO_EMAIL} LIKE '%%tuspeaking%%' OR p.id IS NULL)
ORDER BY ca.acuity_datetime, ca.acuity_type, p.{CAMPO_DURATION} DESC
"""
cursor.execute(sql_zoom_detalle, (fecha_inicio, fecha_fin))
zoom_detalle = cursor.fetchall()
print(f"Registros Zoom detallados: {len(zoom_detalle)}")

# Organizar detalle por mes
zoom_por_mes = defaultdict(list)
for reg in zoom_detalle:
    mes = str(reg['fecha'])[:7]
    zoom_por_mes[mes].append(reg)

# Organizar clases por grupo y mes
grupos_clases = defaultdict(lambda: defaultdict(list))
total_sustituciones = 0
total_co_profesor = 0

for clase in clases_raw:
    grupo = clase['grupo']
    mes = str(clase['fecha'])[:7]
    profesor_asignado = clase['profesor_asignado'] or ''
    profesores_presentes = clase['profesores_presentes'] or ''
    num_profesores = clase['num_profesores'] or 0
    
    hay_profesor = num_profesores > 0
    es_sustituto = False
    hay_co_profesor = False
    otro_profesor = ''
    
    if hay_profesor and profesores_presentes:
        lista_profes = [p.strip() for p in profesores_presentes.split(',')]
        asignado_presente = any(es_mismo_profesor(profesor_asignado, p) for p in lista_profes)
        
        if not asignado_presente and lista_profes:
            es_sustituto = True
            otro_profesor = lista_profes[0]
            total_sustituciones += 1
        elif num_profesores > 1:
            hay_co_profesor = True
            otros = [p for p in lista_profes if not es_mismo_profesor(profesor_asignado, p)]
            if otros:
                otro_profesor = ', '.join(otros)
                total_co_profesor += 1
    
    grupos_clases[grupo][mes].append({
        'fecha': clase['fecha'],
        'profesor_asignado': profesor_asignado,
        'profesores_presentes': profesores_presentes,
        'hay_profesor': hay_profesor,
        'es_sustituto': es_sustituto,
        'hay_co_profesor': hay_co_profesor,
        'otro_profesor': otro_profesor,
        'meeting_id': clase['zoom_meetingid']
    })

print(f"Total clases: {len(clases_raw)}")
print(f"Sustituciones: {total_sustituciones}, Co-profesores: {total_co_profesor}")

# Crear Excel
wb = Workbook()
wb.remove(wb.active)

# Estilos
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
profesor_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')

# ========== HOJAS DE RESUMEN POR MES ==========
for mes in meses:
    mes_num = mes[5:7]
    mes_nombre = meses_es_largo.get(mes_num, mes)
    year = mes[:4]
    
    ws = wb.create_sheet(f"{mes_nombre} {year}")
    
    max_clases = 0
    for grupo in grupos_clases:
        if mes in grupos_clases[grupo]:
            max_clases = max(max_clases, len(grupos_clases[grupo][mes]))
    
    if max_clases == 0:
        continue
    
    # Leyenda
    ws['A1'] = 'Leyenda:'
    ws['B1'] = '✓ OK'
    ws['B1'].fill = green_fill
    ws['C1'] = '⟳ Sustituto'
    ws['C1'].fill = yellow_fill
    ws['D1'] = '+ Co-prof'
    ws['D1'].fill = blue_fill
    ws['E1'] = '✗ Sin prof'
    ws['E1'].fill = red_fill
    
    # Cabecera
    row = 3
    headers = ['#', 'Grupo', 'Profesor Asignado', 'Alum'] + [f'C{i+1}' for i in range(max_clases)] + ['Tot', 'OK', 'Sus', 'Co', '%']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center
    
    row = 4
    grupo_num = 1
    total_clases_mes = 0
    total_ok_mes = 0
    total_sust_mes = 0
    total_co_mes = 0
    
    for grupo in sorted(grupos_clases.keys()):
        if mes not in grupos_clases[grupo]:
            continue
        
        clases_grupo = grupos_clases[grupo][mes]
        if not clases_grupo:
            continue
            
        total_grupo = len(clases_grupo)
        ok_grupo = sum(1 for c in clases_grupo if c['hay_profesor'] and not c['es_sustituto'] and not c['hay_co_profesor'])
        sust_grupo = sum(1 for c in clases_grupo if c['es_sustituto'])
        co_grupo = sum(1 for c in clases_grupo if c['hay_co_profesor'])
        
        total_clases_mes += total_grupo
        total_ok_mes += ok_grupo
        total_sust_mes += sust_grupo
        total_co_mes += co_grupo
        
        info = grupos_info.get(grupo, {'profesor': 'Desconocido', 'alumnos': 0})
        
        ws.cell(row=row, column=1, value=grupo_num).border = thin_border
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=grupo).border = thin_border
        cell_prof = ws.cell(row=row, column=3, value=info['profesor'])
        cell_prof.border = thin_border
        cell_prof.fill = profesor_fill
        ws.cell(row=row, column=4, value=info['alumnos']).border = thin_border
        ws.cell(row=row, column=4).alignment = center
        
        for i, clase in enumerate(clases_grupo):
            col = 5 + i
            fecha_str = clase['fecha'].strftime('%d/%m') if hasattr(clase['fecha'], 'strftime') else str(clase['fecha'])[5:10]
            cell = ws.cell(row=row, column=col, value=fecha_str)
            cell.border = thin_border
            cell.alignment = center
            
            if clase['hay_profesor']:
                if clase['hay_co_profesor']:
                    cell.fill = blue_fill
                    cell.comment = Comment(f"Co-profesor: {clase['otro_profesor']}", "Sistema")
                elif clase['es_sustituto']:
                    cell.fill = yellow_fill
                    cell.comment = Comment(f"Sustituto: {clase['otro_profesor']}", "Sistema")
                else:
                    cell.fill = green_fill
            else:
                cell.fill = red_fill
        
        for i in range(len(clases_grupo), max_clases):
            col = 5 + i
            ws.cell(row=row, column=col, value='').border = thin_border
        
        col_total = 5 + max_clases
        ws.cell(row=row, column=col_total, value=total_grupo).border = thin_border
        ws.cell(row=row, column=col_total).alignment = center
        ws.cell(row=row, column=col_total + 1, value=ok_grupo).border = thin_border
        ws.cell(row=row, column=col_total + 1).alignment = center
        ws.cell(row=row, column=col_total + 2, value=sust_grupo if sust_grupo > 0 else '-').border = thin_border
        ws.cell(row=row, column=col_total + 2).alignment = center
        ws.cell(row=row, column=col_total + 3, value=co_grupo if co_grupo > 0 else '-').border = thin_border
        ws.cell(row=row, column=col_total + 3).alignment = center
        
        pct = round(100 * (ok_grupo + sust_grupo + co_grupo) / total_grupo) if total_grupo > 0 else 0
        cell_pct = ws.cell(row=row, column=col_total + 4, value=f"{pct}%")
        cell_pct.border = thin_border
        cell_pct.alignment = center
        if pct == 100:
            cell_pct.fill = green_fill
        elif pct >= 80:
            cell_pct.fill = yellow_fill
        else:
            cell_pct.fill = red_fill
        
        row += 1
        grupo_num += 1
    
    # Totales
    for col in range(1, 5):
        ws.cell(row=row, column=col, value='').border = thin_border
        ws.cell(row=row, column=col).fill = total_fill
    ws.cell(row=row, column=2, value='TOTAL').font = Font(bold=True)
    
    for col in range(5, 5 + max_clases):
        ws.cell(row=row, column=col, value='').border = thin_border
        ws.cell(row=row, column=col).fill = total_fill
    
    col_total = 5 + max_clases
    for i, val in enumerate([total_clases_mes, total_ok_mes, total_sust_mes, total_co_mes]):
        cell = ws.cell(row=row, column=col_total + i, value=val)
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.fill = total_fill
    
    pct_mes = round(100 * (total_ok_mes + total_sust_mes + total_co_mes) / total_clases_mes) if total_clases_mes > 0 else 0
    cell = ws.cell(row=row, column=col_total + 4, value=f"{pct_mes}%")
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.alignment = center
    cell.fill = total_fill
    
    print(f"  {mes_nombre} {year}: OK={total_ok_mes}, Sust={total_sust_mes}, Co={total_co_mes}, Sin={total_clases_mes-total_ok_mes-total_sust_mes-total_co_mes} ({pct_mes}%)")
    
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 5
    for i in range(max_clases):
        ws.column_dimensions[get_column_letter(5 + i)].width = 6
    for i in range(5):
        ws.column_dimensions[get_column_letter(5 + max_clases + i)].width = 4
    ws.freeze_panes = 'E4'

# ========== HOJAS DE DETALLE ZOOM POR MES ==========
print("\nGenerando hojas de detalle Zoom...")
for mes in meses:
    mes_num = mes[5:7]
    mes_nombre = meses_es.get(mes_num, mes)
    year = mes[:4]
    
    registros = zoom_por_mes.get(mes, [])
    if not registros:
        continue
    
    ws = wb.create_sheet(f"Zoom {mes_nombre} {year}")
    
    # Cabecera
    headers = ['Fecha', 'Hora', 'Grupo', 'Profesor Asignado', 'Meeting ID', 'Participante', 'Email', 'Rol', 'Duración (min)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center
    
    row = 2
    for reg in registros:
        fecha_str = reg['fecha'].strftime('%d/%m/%Y') if hasattr(reg['fecha'], 'strftime') else str(reg['fecha'])
        hora_str = str(reg['hora'])[:5] if reg['hora'] else ''
        
        ws.cell(row=row, column=1, value=fecha_str).border = thin_border
        ws.cell(row=row, column=2, value=hora_str).border = thin_border
        ws.cell(row=row, column=3, value=reg['grupo']).border = thin_border
        ws.cell(row=row, column=4, value=reg['profesor_asignado']).border = thin_border
        ws.cell(row=row, column=5, value=reg['zoom_meetingid']).border = thin_border
        
        participante = reg['participant_name'] or '-'
        if '(Host)' in participante or '(Co-host)' in participante:
            participante = participante.replace(' (Host)', '').replace(' (Co-host)', '')
        
        cell_part = ws.cell(row=row, column=6, value=participante)
        cell_part.border = thin_border
        
        ws.cell(row=row, column=7, value=reg['participant_email'] or '-').border = thin_border
        
        rol_cell = ws.cell(row=row, column=8, value=reg['rol'] if reg['participant_name'] else '-')
        rol_cell.border = thin_border
        rol_cell.alignment = center
        if reg['rol'] == 'Host':
            rol_cell.fill = green_fill
        elif reg['rol'] == 'Co-host':
            rol_cell.fill = blue_fill
        
        dur_cell = ws.cell(row=row, column=9, value=reg['duracion_participante'] if reg['duracion_participante'] else '-')
        dur_cell.border = thin_border
        dur_cell.alignment = center
        
        row += 1
    
    print(f"  Zoom {mes_nombre} {year}: {len(registros)} registros")
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.freeze_panes = 'A2'


# ============================================
# HOJAS DE PUNTUALIDAD
# ============================================
print("\nGenerando hojas de puntualidad...")

# Consulta de puntualidad
sql_puntualidad = f"""
SELECT 
    ca.acuity_type as grupo,
    ca.acuity_calendar as profesor_asignado,
    DATE(ca.acuity_datetime) as fecha,
    TIME(ca.acuity_datetime) as hora_programada,
    ca.acuity_duration as duracion_programada,
    p.{CAMPO_NAME},
    p.{CAMPO_EMAIL},
    TIME(p.join_time) as hora_conexion,
    p.{CAMPO_DURATION} as duracion_real,
    TIMESTAMPDIFF(MINUTE, TIME(ca.acuity_datetime), TIME(p.join_time)) as minutos_tarde
FROM {TABLA_CITAS} ca
JOIN {TABLA_PART} p ON p.zoom_meetingid = ca.zoom_meetingid
WHERE ca.acuity_datetime >= %s 
  AND ca.acuity_datetime <= %s
  AND ca.acuity_firstname != 'Booking'
  AND p.{CAMPO_EMAIL} LIKE '%%tuspeaking%%'
  AND p.{CAMPO_DURATION} >= 30
GROUP BY ca.zoom_meetingid, p.{CAMPO_EMAIL}
ORDER BY ca.acuity_calendar, ca.acuity_datetime
"""
cursor.execute(sql_puntualidad, (fecha_inicio, fecha_fin))
puntualidad_data = cursor.fetchall()
print(f"Registros de puntualidad: {len(puntualidad_data)}")

# Profesores en Sudáfrica (1 hora de diferencia con España)
profesores_sudafrica = ['amber', 'candice', 'kate', 'steve', 'rj', 'jessica', 'reddy', 'klopper', 'gendron', 'lazenby', 'bedford', 'marchant', 'josephs']

# Procesar datos por profesor
prof_stats = defaultdict(lambda: {
    'clases': 0,
    'puntual': 0,
    'tarde': 0,
    'muy_tarde': 0,
    'total_minutos_tarde': 0,
    'total_duracion': 0,
    'duracion_corta': 0
})

# Lista para detalle con ajuste de zona horaria
puntualidad_ajustada = []

for reg in puntualidad_data:
    prof = reg['profesor_asignado'] or 'Desconocido'
    minutos_tarde_orig = reg['minutos_tarde'] or 0
    duracion = reg['duracion_real'] or 0
    
    # Ajustar zona horaria para profesores de Sudáfrica
    prof_lower = prof.lower()
    es_sudafrica = any(p in prof_lower for p in profesores_sudafrica)
    minutos_tarde = minutos_tarde_orig
    if es_sudafrica and minutos_tarde <= -50 and minutos_tarde >= -70:
        minutos_tarde = minutos_tarde + 60
    elif es_sudafrica and minutos_tarde >= 50 and minutos_tarde <= 70:
        minutos_tarde = minutos_tarde - 60
    
    # Guardar para detalle
    reg_copia = dict(reg)
    reg_copia['minutos_tarde_ajustado'] = minutos_tarde
    puntualidad_ajustada.append(reg_copia)
    
    prof_stats[prof]['clases'] += 1
    prof_stats[prof]['total_minutos_tarde'] += minutos_tarde
    prof_stats[prof]['total_duracion'] += duracion
    
    if minutos_tarde <= 2 and minutos_tarde >= -5:
        prof_stats[prof]['puntual'] += 1
    elif minutos_tarde <= 5:
        prof_stats[prof]['tarde'] += 1
    else:
        prof_stats[prof]['muy_tarde'] += 1
    
    if duracion < 55:
        prof_stats[prof]['duracion_corta'] += 1

# Hoja Puntualidad Resumen
ws_punt_res = wb.create_sheet("Puntualidad Resumen")

headers_punt = ['Profesor', 'Clases', 'Puntual (±2min)', '%', 'Tarde (3-5min)', '%', 
                'Muy tarde (>5min)', '%', 'Promedio tarde', 'Duración media', 'Corta (<55min)', '%']
for col, header in enumerate(headers_punt, 1):
    cell = ws_punt_res.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.border = thin_border
    cell.alignment = center

row = 2
for prof in sorted(prof_stats.keys()):
    stats = prof_stats[prof]
    clases = stats['clases']
    if clases == 0:
        continue
    
    pct_puntual = (stats['puntual'] / clases * 100) if clases > 0 else 0
    pct_tarde = (stats['tarde'] / clases * 100) if clases > 0 else 0
    pct_muy_tarde = (stats['muy_tarde'] / clases * 100) if clases > 0 else 0
    prom_tarde = stats['total_minutos_tarde'] / clases if clases > 0 else 0
    dur_media = stats['total_duracion'] / clases if clases > 0 else 0
    pct_corta = (stats['duracion_corta'] / clases * 100) if clases > 0 else 0
    
    ws_punt_res.cell(row=row, column=1, value=prof).border = thin_border
    ws_punt_res.cell(row=row, column=2, value=clases).border = thin_border
    ws_punt_res.cell(row=row, column=3, value=stats['puntual']).border = thin_border
    
    pct_cell = ws_punt_res.cell(row=row, column=4, value=pct_puntual/100)
    pct_cell.number_format = '0%'
    pct_cell.border = thin_border
    if pct_puntual >= 90:
        pct_cell.fill = green_fill
    elif pct_puntual >= 75:
        pct_cell.fill = yellow_fill
    else:
        pct_cell.fill = red_fill
    
    ws_punt_res.cell(row=row, column=5, value=stats['tarde']).border = thin_border
    pct_cell2 = ws_punt_res.cell(row=row, column=6, value=pct_tarde/100)
    pct_cell2.number_format = '0%'
    pct_cell2.border = thin_border
    
    ws_punt_res.cell(row=row, column=7, value=stats['muy_tarde']).border = thin_border
    pct_cell3 = ws_punt_res.cell(row=row, column=8, value=pct_muy_tarde/100)
    pct_cell3.number_format = '0%'
    pct_cell3.border = thin_border
    if pct_muy_tarde > 10:
        pct_cell3.fill = red_fill
    elif pct_muy_tarde > 5:
        pct_cell3.fill = yellow_fill
    
    prom_cell = ws_punt_res.cell(row=row, column=9, value=round(prom_tarde, 1))
    prom_cell.border = thin_border
    if prom_tarde > 3:
        prom_cell.fill = red_fill
    elif prom_tarde > 2:
        prom_cell.fill = yellow_fill
    
    dur_cell = ws_punt_res.cell(row=row, column=10, value=round(dur_media, 0))
    dur_cell.border = thin_border
    if dur_media < 55:
        dur_cell.fill = red_fill
    elif dur_media < 58:
        dur_cell.fill = yellow_fill
    
    ws_punt_res.cell(row=row, column=11, value=stats['duracion_corta']).border = thin_border
    pct_cell4 = ws_punt_res.cell(row=row, column=12, value=pct_corta/100)
    pct_cell4.number_format = '0%'
    pct_cell4.border = thin_border
    if pct_corta > 15:
        pct_cell4.fill = red_fill
    elif pct_corta > 10:
        pct_cell4.fill = yellow_fill
    
    row += 1

for col in range(1, 13):
    ws_punt_res.column_dimensions[chr(64+col)].width = 15
ws_punt_res.column_dimensions['A'].width = 25
ws_punt_res.freeze_panes = 'A2'

print(f"  Puntualidad Resumen: {len(prof_stats)} profesores")

# Hoja Puntualidad Detalle
ws_punt_det = wb.create_sheet("Puntualidad Detalle")

headers_det = ['Fecha', 'Hora Prog.', 'Grupo', 'Profesor', 'Hora Conexión', 
               'Tarde (min)', 'Dur. Prog.', 'Dur. Real', 'Diferencia']
for col, header in enumerate(headers_det, 1):
    cell = ws_punt_det.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.border = thin_border
    cell.alignment = center

row = 2
for reg in puntualidad_ajustada:
    fecha = reg['fecha']
    hora_prog = str(reg['hora_programada'])[:5] if reg['hora_programada'] else '-'
    hora_con = str(reg['hora_conexion'])[:5] if reg['hora_conexion'] else '-'
    minutos_tarde = reg['minutos_tarde_ajustado']
    dur_prog = reg['duracion_programada'] or 60
    dur_real = reg['duracion_real'] or 0
    diferencia = dur_real - dur_prog
    
    ws_punt_det.cell(row=row, column=1, value=fecha).border = thin_border
    ws_punt_det.cell(row=row, column=2, value=hora_prog).border = thin_border
    ws_punt_det.cell(row=row, column=3, value=reg['grupo']).border = thin_border
    ws_punt_det.cell(row=row, column=4, value=reg['profesor_asignado']).border = thin_border
    ws_punt_det.cell(row=row, column=5, value=hora_con).border = thin_border
    
    tarde_cell = ws_punt_det.cell(row=row, column=6, value=minutos_tarde)
    tarde_cell.border = thin_border
    tarde_cell.alignment = center
    if minutos_tarde > 5:
        tarde_cell.fill = red_fill
    elif minutos_tarde > 2:
        tarde_cell.fill = yellow_fill
    elif minutos_tarde >= -5 and minutos_tarde <= 2:
        tarde_cell.fill = green_fill
    
    ws_punt_det.cell(row=row, column=7, value=dur_prog).border = thin_border
    ws_punt_det.cell(row=row, column=8, value=dur_real).border = thin_border
    
    dif_cell = ws_punt_det.cell(row=row, column=9, value=diferencia)
    dif_cell.border = thin_border
    dif_cell.alignment = center
    if diferencia < -10:
        dif_cell.fill = red_fill
    elif diferencia < -5:
        dif_cell.fill = yellow_fill
    
    row += 1

ws_punt_det.column_dimensions['A'].width = 12
ws_punt_det.column_dimensions['B'].width = 10
ws_punt_det.column_dimensions['C'].width = 45
ws_punt_det.column_dimensions['D'].width = 20
ws_punt_det.column_dimensions['E'].width = 12
ws_punt_det.column_dimensions['F'].width = 12
ws_punt_det.column_dimensions['G'].width = 10
ws_punt_det.column_dimensions['H'].width = 10
ws_punt_det.column_dimensions['I'].width = 10
ws_punt_det.freeze_panes = 'A2'

print(f"  Puntualidad Detalle: {len(puntualidad_ajustada)} registros")

# Guardar
filename = f"Reporte_Profesores_CESCE_{fecha_inicio}_a_{fecha_fin}_formato.xlsx"
filepath = f"/home/aulatuspeaking/www/app/moodle/reportes_cesce/{filename}"
wb.save(filepath)

print(f"\n✅ Reporte generado: {filepath}")

cursor.close()
conn.close()
