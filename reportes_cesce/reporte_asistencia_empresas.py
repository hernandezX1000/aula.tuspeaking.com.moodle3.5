#!/usr/bin/env python3
"""
REPORTE ASISTENCIA EMPRESAS - MULTI-EMPRESA
Lee de mdl_i3code_acuityZoom para todas las empresas excepto CESCE
Verde >= 75%
"""

import sys
import mysql.connector
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
from dateutil.relativedelta import relativedelta
from dateutil.parser import parse

empresa = sys.argv[1] if len(sys.argv) > 1 else 'TODAS'
fecha_inicio = sys.argv[2] if len(sys.argv) > 2 else '2025-09-01'
fecha_fin = sys.argv[3] if len(sys.argv) > 3 else datetime.now().strftime('%Y-%m-%d')

EMPRESAS_PATRONES = {
    'GKN': ["acuity_type LIKE '%GKN%'"],
    'TEKIA': ["acuity_type LIKE '%Tekia%'", "acuity_type LIKE '%TEKIA%'"],
    'TORRES_CARRERA': ["acuity_type LIKE '%Torres%'", "acuity_type LIKE '%Carrera%'"],
    'BABEL': ["acuity_type LIKE '%Babel%'"],
    'LIN3S': ["acuity_type LIKE '%Lin3s%'"],
    'SODENA': ["acuity_type LIKE '%Sodena%'"],
    'BEMOBILE': ["acuity_type LIKE '%Bemobile%'"],
    'SERVIGUIDE': ["acuity_type LIKE '%Serviguide%'"],
    'NETTRIM': ["acuity_type LIKE '%Nettrim%'", "acuity_type LIKE '%Attrim%'"],
    'E2Y': ["acuity_type LIKE '%E2Y%'"],
    'OHMYBOX': ["acuity_type LIKE '%Ohmybox%'", "acuity_type LIKE '%Oh My Box%'"]
}

def get_filtro_empresa(emp):
    if emp == 'TODAS':
        all_patterns = []
        for patterns in EMPRESAS_PATRONES.values():
            all_patterns.extend(patterns)
        return "(" + " OR ".join(all_patterns) + ")"
    elif emp in EMPRESAS_PATRONES:
        patterns = EMPRESAS_PATRONES[emp]
        return "(" + " OR ".join(patterns) + ")"
    else:
        return "1=0"

print("=" * 60)
print("  REPORTE ASISTENCIA EMPRESAS")
print(f"  Empresa: {empresa}")
print(f"  Periodo: {fecha_inicio} a {fecha_fin}")
print(f"  Fecha generacion: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 60)

conn = mysql.connector.connect(
    host='localhost', database='aulatuspeaking35',
    user='moodle35', password='TuspeakingFix2025!'
)
cursor = conn.cursor(dictionary=True)

inicio = parse(fecha_inicio)
fin = parse(fecha_fin)
meses = []
current = inicio.replace(day=1)
while current <= fin:
    meses.append(current.strftime('%Y-%m'))
    current += relativedelta(months=1)

meses_es = {'01': 'Ene', '02': 'Feb', '03': 'Mar', '04': 'Abr', '05': 'May', '06': 'Jun',
            '07': 'Jul', '08': 'Ago', '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dic'}

print(f"Meses: {', '.join(meses)}")

filtro_empresa = get_filtro_empresa(empresa)

sql = f"""
SELECT 
    acuity_firstname as nombre,
    acuity_lastname as apellidos,
    acuity_email as email,
    SUBSTRING(acuity_datetime, 1, 7) as mes,
    COUNT(*) as programadas,
    SUM(CASE WHEN zoom_clasecompletada = 1 THEN 1 ELSE 0 END) as asistidas
FROM mdl_i3code_acuityZoom
WHERE acuity_datetime >= %s 
  AND acuity_datetime <= %s 
  AND acuity_firstname != 'Booking'
  AND acuity_email IS NOT NULL
  AND acuity_email != ''
  AND {filtro_empresa}
GROUP BY acuity_email, acuity_firstname, acuity_lastname, SUBSTRING(acuity_datetime, 1, 7)
ORDER BY acuity_lastname, acuity_firstname, mes
"""
cursor.execute(sql, (fecha_inicio, fecha_fin))
alumnos = {}
for row in cursor.fetchall():
    email = row['email'].lower() if row['email'] else 'sin_email'
    if email not in alumnos:
        alumnos[email] = {'nombre': row['nombre'] or '', 'apellidos': row['apellidos'] or '', 'meses': {}}
    alumnos[email]['meses'][row['mes']] = {'prog': row['programadas'], 'asist': row['asistidas']}
print(f"Alumnos: {len(alumnos)}")

sql_grupos = f"""
SELECT 
    acuity_type as grupo,
    acuity_firstname as nombre,
    acuity_lastname as apellidos,
    acuity_email as email,
    SUBSTRING(acuity_datetime, 1, 7) as mes,
    COUNT(*) as prog,
    SUM(CASE WHEN zoom_clasecompletada = 1 THEN 1 ELSE 0 END) as asist
FROM mdl_i3code_acuityZoom
WHERE acuity_datetime >= %s 
  AND acuity_datetime <= %s 
  AND acuity_firstname != 'Booking'
  AND {filtro_empresa}
GROUP BY acuity_type, acuity_email, acuity_firstname, acuity_lastname, SUBSTRING(acuity_datetime, 1, 7)
ORDER BY acuity_type, acuity_lastname, mes
"""
cursor.execute(sql_grupos, (fecha_inicio, fecha_fin))
grupos = defaultdict(lambda: defaultdict(lambda: {'nombre': '', 'apellido': '', 'meses': {}}))
for row in cursor.fetchall():
    email = row['email'].lower() if row['email'] else 'sin_email'
    grupos[row['grupo']][email]['nombre'] = row['nombre'] or ''
    grupos[row['grupo']][email]['apellido'] = row['apellidos'] or ''
    grupos[row['grupo']][email]['meses'][row['mes']] = {'prog': row['prog'], 'asist': row['asist']}
print(f"Grupos: {len(grupos)}")

sql_det = f"""
SELECT 
    az.acuity_datetime,
    az.acuity_firstname as nombre,
    az.acuity_lastname as apellidos,
    az.acuity_email as email,
    az.zoom_meetingid,
    az.zoom_clasecompletada,
    az.acuity_type,
    (SELECT GROUP_CONCAT(DISTINCT p.zoom_name ORDER BY p.zoom_duration DESC SEPARATOR ', ')
     FROM mdl_i3code_acuityZoom_participants p 
     WHERE p.zoom_meetingid = az.zoom_meetingid 
     AND p.zoom_duration >= 1200) as participantes
FROM mdl_i3code_acuityZoom az
WHERE az.acuity_datetime >= %s 
  AND az.acuity_datetime <= %s 
  AND az.acuity_firstname != 'Booking'
  AND {filtro_empresa}
ORDER BY az.acuity_datetime, az.acuity_lastname
"""
cursor.execute(sql_det, (fecha_inicio, fecha_fin))
detalle = cursor.fetchall()
print(f"Detalle: {len(detalle)}")

wb = Workbook()
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
subheader_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
grupo_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center')

def get_fill(pct):
    if pct >= 75: return green_fill
    elif pct >= 50: return yellow_fill
    else: return red_fill

def get_estado(pct):
    if pct >= 75: return 'OK'
    elif pct >= 50: return '!!'
    else: return 'X'

# HOJA 1: RESUMEN ASISTENCIA
ws = wb.active
ws.title = "Resumen Asistencia"
ws['A1'] = f'REPORTE DE ASISTENCIA {empresa} ({fecha_inicio} a {fecha_fin})'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:D1')
ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Verde >=75%, Amarillo >=50%, Rojo <50%"

row = 5
for col, h in enumerate(['Nombre', 'Apellidos', 'Email'], 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

col = 4
for mes in meses:
    cell = ws.cell(row=row, column=col, value=meses_es[mes[5:7]])
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
    for c in range(col, col+3):
        ws.cell(row=row, column=c).fill = header_fill
        ws.cell(row=row, column=c).border = thin_border
    col += 3

cell = ws.cell(row=row, column=col, value='TOTAL')
cell.fill = header_fill
cell.font = header_font
ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
for c in range(col, col+3):
    ws.cell(row=row, column=c).fill = header_fill
    ws.cell(row=row, column=c).border = thin_border

row = 6
for c in range(1, 4):
    ws.cell(row=row, column=c).fill = subheader_fill
    ws.cell(row=row, column=c).border = thin_border
col = 4
for _ in meses:
    for label in ['Prev', 'Asist', '%']:
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = center
        cell.font = Font(bold=True, size=9)
        col += 1
for label in ['Prev', 'Asist', '%']:
    cell = ws.cell(row=row, column=col, value=label)
    cell.fill = subheader_fill
    cell.border = thin_border
    cell.alignment = center
    cell.font = Font(bold=True, size=9)
    col += 1

row = 7
for email, emp in sorted(alumnos.items(), key=lambda x: (x[1]['apellidos'], x[1]['nombre'])):
    ws.cell(row=row, column=1, value=emp['nombre']).border = thin_border
    ws.cell(row=row, column=2, value=emp['apellidos']).border = thin_border
    ws.cell(row=row, column=3, value=email).border = thin_border
    
    col = 4
    total_p, total_a = 0, 0
    for mes in meses:
        p = emp['meses'].get(mes, {}).get('prog', 0)
        a = emp['meses'].get(mes, {}).get('asist', 0)
        pct = round(100*a/p) if p > 0 else 0
        ws.cell(row=row, column=col, value=p if p else '-').border = thin_border
        ws.cell(row=row, column=col).alignment = center
        ws.cell(row=row, column=col+1, value=a if a else '-').border = thin_border
        ws.cell(row=row, column=col+1).alignment = center
        c = ws.cell(row=row, column=col+2)
        c.border = thin_border
        c.alignment = center
        if p:
            c.value = pct / 100
            c.number_format = '0%'
            c.fill = get_fill(pct)
        else:
            c.value = '-'
        total_p += p
        total_a += a
        col += 3
    
    pct_total = round(100*total_a/total_p) if total_p > 0 else 0
    ws.cell(row=row, column=col, value=total_p).border = thin_border
    ws.cell(row=row, column=col).alignment = center
    ws.cell(row=row, column=col).fill = total_fill
    ws.cell(row=row, column=col+1, value=total_a).border = thin_border
    ws.cell(row=row, column=col+1).alignment = center
    ws.cell(row=row, column=col+1).fill = total_fill
    c = ws.cell(row=row, column=col+2)
    c.value = pct_total / 100 if total_p else 0
    c.number_format = '0%'
    c.border = thin_border
    c.alignment = center
    c.fill = get_fill(pct_total) if total_p else total_fill
    row += 1

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 30

# HOJA 2: ASISTENCIA POR GRUPOS
ws2 = wb.create_sheet("Asistencia por Grupos")
ws2['A1'] = f'ASISTENCIA POR GRUPOS - {empresa}'
ws2['A1'].font = Font(bold=True, size=14)

row = 3
grupo_num = 1
resumen_grupos = []

for grupo_nombre, alumnos_grupo in sorted(grupos.items()):
    ws2.cell(row=row, column=1, value=f"#{grupo_num}")
    ws2.cell(row=row, column=2, value=grupo_nombre[:60])
    for c in range(1, 5 + len(meses)*3 + 4):
        ws2.cell(row=row, column=c).fill = grupo_fill
        ws2.cell(row=row, column=c).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    headers = ['Nombre', 'Apellido', 'Email']
    for i, h in enumerate(headers, 1):
        cell = ws2.cell(row=row, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    col = 4
    for mes in meses:
        for label in ['P', 'A', '%']:
            cell = ws2.cell(row=row, column=col, value=f"{meses_es[mes[5:7]]} {label}")
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.border = thin_border
            cell.alignment = center
            col += 1
    
    for label in ['Total P', 'Total A', '%', 'Estado']:
        cell = ws2.cell(row=row, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center
        col += 1
    row += 1
    
    grupo_total_p, grupo_total_a = 0, 0
    for email, datos in sorted(alumnos_grupo.items(), key=lambda x: x[1]['apellido']):
        ws2.cell(row=row, column=1, value=datos['nombre']).border = thin_border
        ws2.cell(row=row, column=2, value=datos['apellido']).border = thin_border
        ws2.cell(row=row, column=3, value=email).border = thin_border
        
        col = 4
        alumno_total_p, alumno_total_a = 0, 0
        for mes in meses:
            p = datos['meses'].get(mes, {}).get('prog', 0)
            a = datos['meses'].get(mes, {}).get('asist', 0)
            pct = round(100*a/p) if p > 0 else 0
            
            ws2.cell(row=row, column=col, value=p if p else '-').border = thin_border
            ws2.cell(row=row, column=col).alignment = center
            ws2.cell(row=row, column=col+1, value=a if a else '-').border = thin_border
            ws2.cell(row=row, column=col+1).alignment = center
            c = ws2.cell(row=row, column=col+2)
            c.border = thin_border
            c.alignment = center
            if p:
                c.value = pct / 100
                c.number_format = '0%'
                c.fill = get_fill(pct)
            else:
                c.value = '-'
            
            alumno_total_p += p
            alumno_total_a += a
            col += 3
        
        pct_alumno = round(100*alumno_total_a/alumno_total_p) if alumno_total_p > 0 else 0
        ws2.cell(row=row, column=col, value=alumno_total_p).border = thin_border
        ws2.cell(row=row, column=col).alignment = center
        ws2.cell(row=row, column=col+1, value=alumno_total_a).border = thin_border
        ws2.cell(row=row, column=col+1).alignment = center
        c = ws2.cell(row=row, column=col+2)
        c.value = pct_alumno / 100 if alumno_total_p else 0
        c.number_format = '0%'
        c.border = thin_border
        c.alignment = center
        c.fill = get_fill(pct_alumno)
        ws2.cell(row=row, column=col+3, value=get_estado(pct_alumno)).border = thin_border
        ws2.cell(row=row, column=col+3).alignment = center
        
        grupo_total_p += alumno_total_p
        grupo_total_a += alumno_total_a
        row += 1
    
    pct_grupo = round(100*grupo_total_a/grupo_total_p) if grupo_total_p > 0 else 0
    resumen_grupos.append({'num': grupo_num, 'nombre': grupo_nombre[:50], 'alumnos': len(alumnos_grupo),
                          'prog': grupo_total_p, 'asist': grupo_total_a, 'pct': pct_grupo})
    row += 1
    grupo_num += 1

row += 2
ws2.cell(row=row, column=1, value="RESUMEN GLOBAL POR GRUPO")
ws2.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 1

headers = ['#', 'Grupo', 'Alumnos', 'Programadas', 'Asistidas', '%', 'Estado']
for i, h in enumerate(headers, 1):
    cell = ws2.cell(row=row, column=i, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
row += 1

for rg in resumen_grupos:
    ws2.cell(row=row, column=1, value=rg['num']).border = thin_border
    ws2.cell(row=row, column=2, value=rg['nombre']).border = thin_border
    ws2.cell(row=row, column=3, value=rg['alumnos']).border = thin_border
    ws2.cell(row=row, column=3).alignment = center
    ws2.cell(row=row, column=4, value=rg['prog']).border = thin_border
    ws2.cell(row=row, column=4).alignment = center
    ws2.cell(row=row, column=5, value=rg['asist']).border = thin_border
    ws2.cell(row=row, column=5).alignment = center
    c = ws2.cell(row=row, column=6)
    c.value = rg['pct'] / 100
    c.number_format = '0%'
    c.border = thin_border
    c.alignment = center
    c.fill = get_fill(rg['pct'])
    ws2.cell(row=row, column=7, value=get_estado(rg['pct'])).border = thin_border
    ws2.cell(row=row, column=7).alignment = center
    row += 1

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 50
ws2.column_dimensions['C'].width = 30

# HOJA 3: DETALLE ZOOM
ws3 = wb.create_sheet("Detalle Zoom")
ws3['A1'] = f'DETALLE DE CLASES - {empresa}'
ws3['A1'].font = Font(bold=True, size=14)

headers = ['Fecha', 'Hora', 'Nombre', 'Apellidos', 'Email', 'Meeting ID', 'Estado', 'Participantes', 'Tipo Clase']
row = 3
for col, h in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

estados = {0: 'Sin datos', 1: 'Asistio', 2: 'Ausencia', 3: 'Pendiente'}
row = 4
for d in detalle:
    try:
        dt = parse(d['acuity_datetime'])
        fecha = dt.strftime('%Y-%m-%d')
        hora = dt.strftime('%H:%M')
    except:
        fecha = str(d['acuity_datetime'])[:10]
        hora = str(d['acuity_datetime'])[11:16] if len(str(d['acuity_datetime'])) > 11 else ''
    
    ws3.cell(row=row, column=1, value=fecha).border = thin_border
    ws3.cell(row=row, column=2, value=hora).border = thin_border
    ws3.cell(row=row, column=3, value=d['nombre'] or '').border = thin_border
    ws3.cell(row=row, column=4, value=d['apellidos'] or '').border = thin_border
    ws3.cell(row=row, column=5, value=d['email'] or '').border = thin_border
    ws3.cell(row=row, column=6, value=d['zoom_meetingid'] or '').border = thin_border
    
    estado = estados.get(d['zoom_clasecompletada'], 'Desconocido')
    c = ws3.cell(row=row, column=7, value=estado)
    c.border = thin_border
    if d['zoom_clasecompletada'] == 1:
        c.fill = green_fill
    elif d['zoom_clasecompletada'] == 2:
        c.fill = red_fill
    elif d['zoom_clasecompletada'] == 0:
        c.fill = yellow_fill
    
    ws3.cell(row=row, column=8, value=(d['participantes'] or '')[:100]).border = thin_border
    ws3.cell(row=row, column=9, value=(d['acuity_type'] or '')[:60]).border = thin_border
    row += 1

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 8
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 30
ws3.column_dimensions['F'].width = 15
ws3.column_dimensions['G'].width = 12
ws3.column_dimensions['H'].width = 50
ws3.column_dimensions['I'].width = 40

archivo = f"/home/aulatuspeaking/www/app/moodle/reportes_cesce/Reporte_{empresa}_{fecha_inicio}_a_{fecha_fin}.xlsx"
wb.save(archivo)

print(f"\n  Hoja 1: Resumen Asistencia")
print(f"  Hoja 2: Asistencia por Grupos")
print(f"  Hoja 3: Detalle Zoom")
print(f"\nReporte: {archivo}")
print(f"   Alumnos: {len(alumnos)}, Grupos: {len(grupos)}, Detalle: {len(detalle)}")

cursor.close()
conn.close()
