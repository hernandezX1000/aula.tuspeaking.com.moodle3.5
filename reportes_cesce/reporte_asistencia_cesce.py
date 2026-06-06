#!/usr/bin/env python3
"""
REPORTE ASISTENCIA CESCE - MENSUALIZADO + DETALLE ZOOM + GRUPOS
Verde >= 75%
"""

import sys
import mysql.connector
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
from dateutil.relativedelta import relativedelta
from dateutil.parser import parse

fecha_inicio = sys.argv[1] if len(sys.argv) > 1 else '2025-09-01'
fecha_fin = sys.argv[2] if len(sys.argv) > 2 else '2025-12-19'
# === DETECCIÓN AUTOMÁTICA DE TABLAS ===
# 2025: tablas cesce/coding (importación manual CSV)
# 2026+: tablas i3code (ingesta automática)
if fecha_inicio >= '2026-01-01':
    TABLA_CITAS = 'mdl_i3code_acuityZoom'
    TABLA_PART = 'mdl_i3code_acuityZoom_participants'
    CAMPO_PART_NAME = 'zoom_name'
    CAMPO_PART_DURATION = 'zoom_duration/60'
    print(f"  Usando tablas i3code (2026+)")
else:
    TABLA_CITAS = 'mdl_cesce_acuityZoom'
    TABLA_PART = 'mdl_coding_zoom_participants'
    CAMPO_PART_NAME = 'participant_name'
    CAMPO_PART_DURATION = 'duration_minutes'
    print(f"  Usando tablas cesce (2025)")


print("=" * 60)
print("  REPORTE ASISTENCIA CESCE MENSUALIZADO")
print(f"  Periodo: {fecha_inicio} a {fecha_fin}")
print(f"  Fecha generacion: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 60)

conn = mysql.connector.connect(
    host='localhost', database='aulatuspeaking35',
    user='moodle35', password='TuspeakingFix2025!'
)
cursor = conn.cursor(dictionary=True)

inicio = parse(fecha_inicio)
fin = parse(fecha_fin)
meses = []
current = inicio.replace(day=1)
while current <= fin:
    meses.append(current.strftime('%Y-%m'))
    current += relativedelta(months=1)

meses_es = {'01': 'Ene', '02': 'Feb', '03': 'Mar', '04': 'Abr', '05': 'May', '06': 'Jun',
            '07': 'Jul', '08': 'Ago', '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dic'}

print(f"Meses: {', '.join(meses)}")

sql = f"""
SELECT e.firstname, e.lastname, e.email, e.codigo_empleado,
    SUBSTRING(ca.acuity_datetime, 1, 7) as mes,
    COUNT(*) as programadas,
    SUM(CASE WHEN ca.zoom_clasecompletada = 1 THEN 1 ELSE 0 END) as asistidas
FROM {TABLA_CITAS} ca
INNER JOIN mdl_cesce_empleados e ON LOWER(e.email) = LOWER(ca.acuity_email)
WHERE ca.acuity_datetime >= %s AND ca.acuity_datetime <= %s AND ca.acuity_firstname != 'Booking'
GROUP BY e.email, e.firstname, e.lastname, e.codigo_empleado, SUBSTRING(ca.acuity_datetime, 1, 7)
ORDER BY e.lastname, e.firstname, mes
"""
cursor.execute(sql, (fecha_inicio, fecha_fin))
empleados = {}
for row in cursor.fetchall():
    email = row['email']
    if email not in empleados:
        empleados[email] = {'nombre': row['firstname'], 'apellidos': row['lastname'], 
                           'codigo': row['codigo_empleado'], 'meses': {}}
    empleados[email]['meses'][row['mes']] = {'prog': row['programadas'], 'asist': row['asistidas']}
print(f"Empleados: {len(empleados)}")

sql_grupos = f"""
SELECT ca.acuity_type as grupo, e.firstname, e.lastname, e.email,
    SUBSTRING(ca.acuity_datetime, 1, 7) as mes, COUNT(*) as prog,
    SUM(CASE WHEN ca.zoom_clasecompletada = 1 THEN 1 ELSE 0 END) as asist
FROM {TABLA_CITAS} ca
INNER JOIN mdl_cesce_empleados e ON LOWER(e.email) = LOWER(ca.acuity_email)
WHERE ca.acuity_datetime >= %s AND ca.acuity_datetime <= %s AND ca.acuity_firstname != 'Booking'
GROUP BY ca.acuity_type, e.email, e.firstname, e.lastname, SUBSTRING(ca.acuity_datetime, 1, 7)
ORDER BY ca.acuity_type, e.lastname, mes
"""
cursor.execute(sql_grupos, (fecha_inicio, fecha_fin))
grupos = defaultdict(lambda: defaultdict(lambda: {'nombre': '', 'apellido': '', 'meses': {}}))
for row in cursor.fetchall():
    grupos[row['grupo']][row['email']]['nombre'] = row['firstname']
    grupos[row['grupo']][row['email']]['apellido'] = row['lastname']
    grupos[row['grupo']][row['email']]['meses'][row['mes']] = {'prog': row['prog'], 'asist': row['asist']}
print(f"Grupos: {len(grupos)}")

sql_det = f"""
SELECT ca.acuity_datetime, e.firstname, e.lastname, e.email, e.codigo_empleado,
    ca.zoom_meetingid, ca.zoom_clasecompletada, ca.acuity_type,
    (SELECT GROUP_CONCAT(DISTINCT p.participant_name ORDER BY p.duration_minutes DESC SEPARATOR ', ')
     FROM mdl_coding_zoom_participants p WHERE p.zoom_meetingid = ca.zoom_meetingid AND p.duration_minutes >= 20) as participantes
FROM {TABLA_CITAS} ca
INNER JOIN mdl_cesce_empleados e ON LOWER(e.email) = LOWER(ca.acuity_email)
WHERE ca.acuity_datetime >= %s AND ca.acuity_datetime <= %s AND ca.acuity_firstname != 'Booking'
ORDER BY ca.acuity_datetime, e.lastname
"""
cursor.execute(sql_det, (fecha_inicio, fecha_fin))
detalle = cursor.fetchall()
print(f"Detalle: {len(detalle)}")

wb = Workbook()
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
subheader_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
grupo_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center')

def get_fill(pct):
    if pct >= 75: return green_fill
    elif pct >= 50: return yellow_fill
    else: return red_fill

def get_estado(pct):
    if pct >= 75: return 'OK'
    elif pct >= 50: return '!!'
    else: return 'X'

ws = wb.active
ws.title = "Resumen Asistencia"
ws['A1'] = f'REPORTE DE ASISTENCIA CESCE ({fecha_inicio} a {fecha_fin})'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:D1')
ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Verde >=75%, Amarillo >=50%, Rojo <50%"

row = 5
for col, h in enumerate(['Nombre', 'Apellidos', 'Codigo', 'Email'], 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

col = 5
for mes in meses:
    cell = ws.cell(row=row, column=col, value=meses_es[mes[5:7]])
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
    for c in range(col, col+3):
        ws.cell(row=row, column=c).fill = header_fill
        ws.cell(row=row, column=c).border = thin_border
    col += 3

cell = ws.cell(row=row, column=col, value='TOTAL')
cell.fill = header_fill
cell.font = header_font
ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
for c in range(col, col+3):
    ws.cell(row=row, column=c).fill = header_fill
    ws.cell(row=row, column=c).border = thin_border
last_col = col + 2

row = 6
for c in range(1, 5):
    ws.cell(row=row, column=c).fill = subheader_fill
    ws.cell(row=row, column=c).border = thin_border
col = 5
for _ in meses:
    for label in ['Prev', 'Asist', '%']:
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = center
        cell.font = Font(bold=True, size=9)
        col += 1
for label in ['Prev', 'Asist', '%']:
    cell = ws.cell(row=row, column=col, value=label)
    cell.fill = subheader_fill
    cell.border = thin_border
    cell.alignment = center
    cell.font = Font(bold=True, size=9)
    col += 1

row = 7
for email, emp in sorted(empleados.items(), key=lambda x: (x[1]['apellidos'], x[1]['nombre'])):
    ws.cell(row=row, column=1, value=emp['nombre']).border = thin_border
    ws.cell(row=row, column=2, value=emp['apellidos']).border = thin_border
    ws.cell(row=row, column=3, value=emp['codigo']).border = thin_border
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=4, value=email).border = thin_border
    
    col = 5
    total_p, total_a = 0, 0
    for mes in meses:
        p = emp['meses'].get(mes, {}).get('prog', 0)
        a = emp['meses'].get(mes, {}).get('asist', 0)
        pct = round(100*a/p) if p > 0 else 0
        ws.cell(row=row, column=col, value=p if p else '-').border = thin_border
        ws.cell(row=row, column=col).alignment = center
        ws.cell(row=row, column=col+1, value=a if a else '-').border = thin_border
        ws.cell(row=row, column=col+1).alignment = center
        c = ws.cell(row=row, column=col+2)
        c.border = thin_border
        c.alignment = center
        if p:
            c.value = pct / 100
            c.number_format = '0%'
            c.fill = get_fill(pct)
        else:
            c.value = '-'
        col += 3
        total_p += p
        total_a += a
    
    pct_t = round(100*total_a/total_p) if total_p > 0 else 0
    ws.cell(row=row, column=col, value=total_p).border = thin_border
    ws.cell(row=row, column=col).alignment = center
    ws.cell(row=row, column=col+1, value=total_a).border = thin_border
    ws.cell(row=row, column=col+1).alignment = center
    c = ws.cell(row=row, column=col+2)
    c.value = pct_t / 100
    c.number_format = '0%'
    c.border = thin_border
    c.alignment = center
    c.fill = get_fill(pct_t)
    row += 1

ws.cell(row=row, column=1, value='TOTALES').font = Font(bold=True)
for c in range(1, 5):
    ws.cell(row=row, column=c).fill = total_fill
    ws.cell(row=row, column=c).border = thin_border

col = 5
gp, ga = 0, 0
for mes in meses:
    mp = sum(e['meses'].get(mes, {}).get('prog', 0) for e in empleados.values())
    ma = sum(e['meses'].get(mes, {}).get('asist', 0) for e in empleados.values())
    mpct = round(100*ma/mp) if mp > 0 else 0
    c = ws.cell(row=row, column=col, value=mp)
    c.fill = total_fill
    c.border = thin_border
    c.alignment = center
    c.font = Font(bold=True)
    col += 1
    c = ws.cell(row=row, column=col, value=ma)
    c.fill = total_fill
    c.border = thin_border
    c.alignment = center
    c.font = Font(bold=True)
    col += 1
    c = ws.cell(row=row, column=col)
    c.value = mpct / 100
    c.number_format = '0%'
    c.fill = total_fill
    c.border = thin_border
    c.alignment = center
    c.font = Font(bold=True)
    col += 1
    gp += mp
    ga += ma

gpct = round(100*ga/gp) if gp > 0 else 0
c = ws.cell(row=row, column=col, value=gp)
c.fill = total_fill
c.border = thin_border
c.alignment = center
c.font = Font(bold=True)
col += 1
c = ws.cell(row=row, column=col, value=ga)
c.fill = total_fill
c.border = thin_border
c.alignment = center
c.font = Font(bold=True)
col += 1
c = ws.cell(row=row, column=col)
c.value = gpct / 100
c.number_format = '0%'
c.fill = total_fill
c.border = thin_border
c.alignment = center
c.font = Font(bold=True)

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 28
for c in range(5, last_col+1):
    ws.column_dimensions[get_column_letter(c)].width = 6
ws.freeze_panes = 'E7'
print("  Hoja 1: Resumen Asistencia")

ws2 = wb.create_sheet("Asistencia por Grupos")
row = 1
ws2.cell(row=row, column=1, value=f"ASISTENCIA POR GRUPOS CESCE ({fecha_inicio} a {fecha_fin})")
ws2.cell(row=row, column=1).font = Font(bold=True, size=14)
row += 2

resumen_grupos = []
for grupo in sorted(grupos.keys()):
    alumnos = grupos[grupo]
    tot_p = sum(sum(a['meses'][m]['prog'] for m in a['meses']) for a in alumnos.values())
    tot_a = sum(sum(a['meses'][m]['asist'] for m in a['meses']) for a in alumnos.values())
    pct_g = round(100*tot_a/tot_p) if tot_p > 0 else 0
    estado = get_estado(pct_g)
    
    resumen_grupos.append({'grupo': grupo, 'alumnos': len(alumnos), 'pct': pct_g})
    
    ws2.cell(row=row, column=1, value=f"{grupo} - Promedio: {pct_g}% {estado}")
    ws2.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    num_cols = 4 + len(meses)*3 + 4
    for c in range(1, num_cols):
        ws2.cell(row=row, column=c).fill = grupo_fill
        ws2.cell(row=row, column=c).border = thin_border
    row += 1
    
    headers = ['Nombre', 'Apellido']
    for mes in meses:
        headers.extend([f'{meses_es[mes[5:7]]} P', f'{meses_es[mes[5:7]]} A', f'{meses_es[mes[5:7]]} %'])
    headers.extend(['Total P', 'Total A', 'Total %', 'Estado'])
    
    for col_idx, h in enumerate(headers, 1):
        c = ws2.cell(row=row, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.border = thin_border
        c.alignment = center
    row += 1
    
    for email in sorted(alumnos.keys(), key=lambda e: (alumnos[e]['apellido'], alumnos[e]['nombre'])):
        a = alumnos[email]
        ws2.cell(row=row, column=1, value=a['nombre']).border = thin_border
        ws2.cell(row=row, column=2, value=a['apellido']).border = thin_border
        
        col = 3
        tp, ta = 0, 0
        for mes in meses:
            p = a['meses'].get(mes, {}).get('prog', 0)
            asist = a['meses'].get(mes, {}).get('asist', 0)
            pct = round(100*asist/p) if p > 0 else 0
            ws2.cell(row=row, column=col, value=p if p else '-').border = thin_border
            ws2.cell(row=row, column=col).alignment = center
            ws2.cell(row=row, column=col+1, value=asist if asist else '-').border = thin_border
            ws2.cell(row=row, column=col+1).alignment = center
            c = ws2.cell(row=row, column=col+2)
            c.border = thin_border
            c.alignment = center
            if p:
                c.value = pct / 100
                c.number_format = '0%'
                c.fill = get_fill(pct)
            else:
                c.value = '-'
            col += 3
            tp += p
            ta += asist
        
        pct_t = round(100*ta/tp) if tp > 0 else 0
        ws2.cell(row=row, column=col, value=tp).border = thin_border
        ws2.cell(row=row, column=col).alignment = center
        ws2.cell(row=row, column=col+1, value=ta).border = thin_border
        ws2.cell(row=row, column=col+1).alignment = center
        c = ws2.cell(row=row, column=col+2)
        c.value = pct_t / 100
        c.number_format = '0%'
        c.border = thin_border
        c.alignment = center
        c.fill = get_fill(pct_t)
        c = ws2.cell(row=row, column=col+3, value=get_estado(pct_t))
        c.border = thin_border
        c.alignment = center
        c.fill = get_fill(pct_t)
        row += 1
    row += 1

row += 1
ws2.cell(row=row, column=1, value="RESUMEN GLOBAL")
ws2.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
for c in range(1, 5):
    ws2.cell(row=row, column=c).fill = grupo_fill
    ws2.cell(row=row, column=c).border = thin_border
row += 1

for col_idx, h in enumerate(['Grupo', 'Alumnos', 'Promedio', 'Estado'], 1):
    c = ws2.cell(row=row, column=col_idx, value=h)
    c.fill = header_fill
    c.font = header_font
    c.border = thin_border
row += 1

for r in resumen_grupos:
    ws2.cell(row=row, column=1, value=r['grupo'][:55]).border = thin_border
    ws2.cell(row=row, column=2, value=r['alumnos']).border = thin_border
    ws2.cell(row=row, column=2).alignment = center
    c = ws2.cell(row=row, column=3)
    c.value = r['pct'] / 100
    c.number_format = '0%'
    c.border = thin_border
    c.alignment = center
    c.fill = get_fill(r['pct'])
    c = ws2.cell(row=row, column=4, value=get_estado(r['pct']))
    c.border = thin_border
    c.alignment = center
    c.fill = get_fill(r['pct'])
    row += 1

ws2.column_dimensions['A'].width = 55
ws2.column_dimensions['B'].width = 20
ws2.freeze_panes = 'C4'
print("  Hoja 2: Asistencia por Grupos")

ws3 = wb.create_sheet("Detalle Zoom")
headers = ['Fecha', 'Hora', 'Nombre', 'Apellidos', 'Email', 'Codigo', 'Meeting ID', 'Estado', 'Participantes', 'Tipo Clase']
for col_idx, h in enumerate(headers, 1):
    c = ws3.cell(row=1, column=col_idx, value=h)
    c.fill = header_fill
    c.font = header_font
    c.border = thin_border

estados = {0: 'Sin datos', 1: 'Asistio', 2: 'Ausencia', 3: 'Pendiente'}
row = 2
for d in detalle:
    fecha = d['acuity_datetime'][:10] if d['acuity_datetime'] else ''
    hora = d['acuity_datetime'][11:16] if d['acuity_datetime'] and len(d['acuity_datetime']) > 11 else ''
    ws3.cell(row=row, column=1, value=fecha).border = thin_border
    ws3.cell(row=row, column=2, value=hora).border = thin_border
    ws3.cell(row=row, column=3, value=d['firstname']).border = thin_border
    ws3.cell(row=row, column=4, value=d['lastname']).border = thin_border
    ws3.cell(row=row, column=5, value=d['email']).border = thin_border
    ws3.cell(row=row, column=6, value=d['codigo_empleado']).border = thin_border
    ws3.cell(row=row, column=7, value=d['zoom_meetingid']).border = thin_border
    c = ws3.cell(row=row, column=8, value=estados.get(d['zoom_clasecompletada'], '?'))
    c.border = thin_border
    c.alignment = center
    if d['zoom_clasecompletada'] == 1: c.fill = green_fill
    elif d['zoom_clasecompletada'] == 2: c.fill = red_fill
    elif d['zoom_clasecompletada'] == 3: c.fill = yellow_fill
    ws3.cell(row=row, column=9, value=d['participantes'] or '-').border = thin_border
    ws3.cell(row=row, column=10, value=d['acuity_type'] or '-').border = thin_border
    row += 1

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 8
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 28
ws3.column_dimensions['F'].width = 10
ws3.column_dimensions['G'].width = 14
ws3.column_dimensions['H'].width = 12
ws3.column_dimensions['I'].width = 50
ws3.column_dimensions['J'].width = 40
ws3.freeze_panes = 'A2'
print("  Hoja 3: Detalle Zoom")

filepath = f"/home/aulatuspeaking/www/app/moodle/reportes_cesce/Reporte_CESCE_{fecha_inicio}_a_{fecha_fin}.xlsx"
wb.save(filepath)
print(f"\nReporte: {filepath}")
print(f"   Empleados: {len(empleados)}, Grupos: {len(grupos)}, Detalle: {len(detalle)}")

cursor.close()
conn.close()
