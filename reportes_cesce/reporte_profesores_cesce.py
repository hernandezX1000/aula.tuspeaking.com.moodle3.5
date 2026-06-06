#!/usr/bin/env python3
"""
REPORTE ASISTENCIA PROFESORES CESCE

Uso: python3 reporte_profesores_cesce.py [fecha_inicio] [fecha_fin]
"""

import sys
import mysql.connector
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta
from dateutil.parser import parse

# Fechas
fecha_inicio = sys.argv[1] if len(sys.argv) > 1 else '2025-09-01'
fecha_fin = sys.argv[2] if len(sys.argv) > 2 else '2025-12-19'

print("=" * 60)
print("  REPORTE ASISTENCIA PROFESORES CESCE")
print(f"  Período: {fecha_inicio} a {fecha_fin}")
print(f"  Fecha generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 60)
print()

# Lista de profesores conocidos
PROFESORES = [
    'rj reddy', 'amber gendron', 'jessica josephs', 'michaela duffell canham',
    'jessica lee bedford', 'odile tezel', 'eldina nasufi', 'candice lazenby',
    'karen malik', 'steve marchant', 'dehlia williams', 'gabor legradi',
    'rachel maccobb', 'daniette horwood'
]

# Conexión BD
conn = mysql.connector.connect(
    host='localhost',
    database='aulatuspeaking35',
    user='moodle35',
    password='TuspeakingFix2025!'
)
cursor = conn.cursor(dictionary=True)

# Obtener meses
inicio = parse(fecha_inicio)
fin = parse(fecha_fin)
meses = []
current = inicio.replace(day=1)
while current <= fin:
    meses.append(current.strftime('%Y-%m'))
    current += relativedelta(months=1)

meses_es = {
    '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
    '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
    '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dic'
}

print(f"Meses incluidos: {', '.join(meses)}")

# Consulta: todas las clases con info del profesor
sql = """
SELECT 
    ca.id,
    ca.acuity_datetime,
    ca.zoom_meetingid,
    ca.acuity_calendar as profesor_acuity,
    ca.acuity_type,
    e.firstname as alumno_nombre,
    e.lastname as alumno_apellido,
    e.email as alumno_email,
    (SELECT GROUP_CONCAT(DISTINCT p.participant_name SEPARATOR ', ')
     FROM mdl_coding_zoom_participants p 
     WHERE p.zoom_meetingid = ca.zoom_meetingid 
     AND (p.participant_name LIKE '%%(Host)%%' OR p.participant_name LIKE '%%(Co-host)%%')
    ) as host_zoom,
    (SELECT GROUP_CONCAT(DISTINCT p.participant_name SEPARATOR ', ')
     FROM mdl_coding_zoom_participants p 
     WHERE p.zoom_meetingid = ca.zoom_meetingid 
     AND p.duration_minutes >= 30
     AND p.participant_name NOT LIKE '%%(Host)%%'
    ) as participantes,
    (SELECT MAX(p.duration_minutes)
     FROM mdl_coding_zoom_participants p 
     WHERE p.zoom_meetingid = ca.zoom_meetingid 
     AND (p.participant_name LIKE '%%(Host)%%' OR p.participant_name LIKE '%%(Co-host)%%')
    ) as duracion_host,
    ca.zoom_clasecompletada
FROM mdl_cesce_acuityZoom ca
LEFT JOIN mdl_cesce_empleados e ON LOWER(e.email) = LOWER(ca.acuity_email)
WHERE ca.acuity_datetime >= %s 
  AND ca.acuity_datetime <= %s
  AND ca.acuity_firstname != 'Booking'
ORDER BY ca.acuity_datetime
"""

cursor.execute(sql, (fecha_inicio, fecha_fin))
clases = cursor.fetchall()

print(f"Clases encontradas: {len(clases)}")

# Analizar profesores
profesores_stats = {}
clases_sin_profesor = []
clases_con_profesor = 0

for clase in clases:
    host = clase['host_zoom'] or ''
    profesor_acuity = clase['profesor_acuity'] or 'Desconocido'
    
    # Extraer nombre del profesor del host
    profesor_nombre = None
    if host:
        # Quitar (Host) o (Co-host)
        nombre_host = host.replace('(Host)', '').replace('(Co-host)', '').strip()
        if ',' in nombre_host:
            nombre_host = nombre_host.split(',')[0].strip()
        profesor_nombre = nombre_host
        clases_con_profesor += 1
    else:
        # Buscar en participantes si hay profesor conocido
        participantes = (clase['participantes'] or '').lower()
        for prof in PROFESORES:
            if prof in participantes:
                profesor_nombre = prof.title()
                clases_con_profesor += 1
                break
    
    if not profesor_nombre:
        clases_sin_profesor.append(clase)
        profesor_nombre = 'SIN PROFESOR'
    
    # Usar profesor de Acuity si no hay host
    key = profesor_nombre if profesor_nombre != 'SIN PROFESOR' else profesor_acuity
    
    if key not in profesores_stats:
        profesores_stats[key] = {
            'nombre': key,
            'clases': 0,
            'con_host': 0,
            'sin_host': 0,
            'meses': {m: {'total': 0, 'con_host': 0} for m in meses}
        }
    
    mes = clase['acuity_datetime'][:7] if clase['acuity_datetime'] else None
    profesores_stats[key]['clases'] += 1
    
    if host or profesor_nombre != 'SIN PROFESOR':
        profesores_stats[key]['con_host'] += 1
        if mes and mes in profesores_stats[key]['meses']:
            profesores_stats[key]['meses'][mes]['con_host'] += 1
    else:
        profesores_stats[key]['sin_host'] += 1
    
    if mes and mes in profesores_stats[key]['meses']:
        profesores_stats[key]['meses'][mes]['total'] += 1

print(f"Clases con profesor identificado: {clases_con_profesor}")
print(f"Clases sin profesor: {len(clases_sin_profesor)}")
print(f"Profesores únicos: {len(profesores_stats)}")

# Crear Excel
wb = Workbook()

# Estilos
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
subheader_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')

# ============ HOJA 1: RESUMEN PROFESORES ============
ws = wb.active
ws.title = "Resumen Profesores"

ws['A1'] = 'REPORTE DE ASISTENCIA PROFESORES CESCE'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:D1')

ws['A2'] = f"Período: {fecha_inicio} a {fecha_fin}"
ws.merge_cells('A2:D2')

ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
ws.merge_cells('A3:D3')

# Cabecera
row = 5
headers = ['Profesor', 'Total Clases', 'Con Host', 'Sin Host', '%']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center

# Añadir columnas por mes
col = 6
for mes in meses:
    mes_num = mes[5:7]
    mes_nombre = meses_es.get(mes_num, mes)
    cell = ws.cell(row=row, column=col, value=mes_nombre)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center
    col += 1

last_col = col - 1

# Datos
row = 6
for prof_key in sorted(profesores_stats.keys()):
    prof = profesores_stats[prof_key]
    col = 1
    
    ws.cell(row=row, column=col, value=prof['nombre']).border = thin_border
    col += 1
    ws.cell(row=row, column=col, value=prof['clases']).border = thin_border
    ws.cell(row=row, column=col).alignment = center
    col += 1
    ws.cell(row=row, column=col, value=prof['con_host']).border = thin_border
    ws.cell(row=row, column=col).alignment = center
    col += 1
    ws.cell(row=row, column=col, value=prof['sin_host']).border = thin_border
    ws.cell(row=row, column=col).alignment = center
    col += 1
    
    pct = round(100 * prof['con_host'] / prof['clases']) if prof['clases'] > 0 else 0
    pct_cell = ws.cell(row=row, column=col, value=f"{pct}%")
    pct_cell.border = thin_border
    pct_cell.alignment = center
    
    if pct >= 90:
        pct_cell.fill = green_fill
    elif pct >= 70:
        pct_cell.fill = yellow_fill
    else:
        pct_cell.fill = red_fill
    col += 1
    
    # Clases por mes
    for mes in meses:
        val = prof['meses'].get(mes, {}).get('total', 0)
        ws.cell(row=row, column=col, value=val if val else '-').border = thin_border
        ws.cell(row=row, column=col).alignment = center
        col += 1
    
    row += 1

# Totales
ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
ws.cell(row=row, column=1).fill = total_fill
ws.cell(row=row, column=1).border = thin_border

total_clases = sum(p['clases'] for p in profesores_stats.values())
total_con = sum(p['con_host'] for p in profesores_stats.values())
total_sin = sum(p['sin_host'] for p in profesores_stats.values())
total_pct = round(100 * total_con / total_clases) if total_clases > 0 else 0

for col, val in enumerate([total_clases, total_con, total_sin, f"{total_pct}%"], 2):
    cell = ws.cell(row=row, column=col, value=val)
    cell.fill = total_fill
    cell.border = thin_border
    cell.alignment = center
    cell.font = Font(bold=True)

# Ajustar anchos
ws.column_dimensions['A'].width = 25
for c in range(2, last_col + 1):
    ws.column_dimensions[get_column_letter(c)].width = 10

ws.freeze_panes = 'B6'

# ============ HOJA 2: CLASES SIN PROFESOR ============
ws2 = wb.create_sheet("Clases Sin Profesor")

headers_sin = ['Fecha', 'Hora', 'Alumno', 'Email', 'Meeting ID', 'Profesor Acuity', 'Participantes']
row = 1
for col, h in enumerate(headers_sin, 1):
    cell = ws2.cell(row=row, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center

row = 2
for clase in clases_sin_profesor:
    fecha = clase['acuity_datetime'][:10] if clase['acuity_datetime'] else ''
    hora = clase['acuity_datetime'][11:16] if clase['acuity_datetime'] and len(clase['acuity_datetime']) > 11 else ''
    
    ws2.cell(row=row, column=1, value=fecha).border = thin_border
    ws2.cell(row=row, column=2, value=hora).border = thin_border
    ws2.cell(row=row, column=3, value=f"{clase['alumno_nombre'] or ''} {clase['alumno_apellido'] or ''}").border = thin_border
    ws2.cell(row=row, column=4, value=clase['alumno_email'] or '').border = thin_border
    ws2.cell(row=row, column=5, value=clase['zoom_meetingid'] or '').border = thin_border
    ws2.cell(row=row, column=6, value=clase['profesor_acuity'] or '').border = thin_border
    ws2.cell(row=row, column=7, value=clase['participantes'] or '-').border = thin_border
    row += 1

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 8
ws2.column_dimensions['C'].width = 25
ws2.column_dimensions['D'].width = 28
ws2.column_dimensions['E'].width = 14
ws2.column_dimensions['F'].width = 25
ws2.column_dimensions['G'].width = 50

ws2.freeze_panes = 'A2'

# ============ HOJA 3: DETALLE TODAS LAS CLASES ============
ws3 = wb.create_sheet("Detalle Clases")

headers_det = ['Fecha', 'Hora', 'Alumno', 'Email Alumno', 'Profesor Acuity', 'Host Zoom', 'Duración Host', 'Meeting ID', 'Estado']
row = 1
for col, h in enumerate(headers_det, 1):
    cell = ws3.cell(row=row, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center

estados = {0: 'Sin datos', 1: 'Asistió', 2: 'Ausencia', 3: 'Pendiente'}

row = 2
for clase in clases:
    fecha = clase['acuity_datetime'][:10] if clase['acuity_datetime'] else ''
    hora = clase['acuity_datetime'][11:16] if clase['acuity_datetime'] and len(clase['acuity_datetime']) > 11 else ''
    
    ws3.cell(row=row, column=1, value=fecha).border = thin_border
    ws3.cell(row=row, column=2, value=hora).border = thin_border
    ws3.cell(row=row, column=3, value=f"{clase['alumno_nombre'] or ''} {clase['alumno_apellido'] or ''}").border = thin_border
    ws3.cell(row=row, column=4, value=clase['alumno_email'] or '').border = thin_border
    ws3.cell(row=row, column=5, value=clase['profesor_acuity'] or '').border = thin_border
    
    host_cell = ws3.cell(row=row, column=6, value=clase['host_zoom'] or '-')
    host_cell.border = thin_border
    if not clase['host_zoom']:
        host_cell.fill = red_fill
    
    ws3.cell(row=row, column=7, value=clase['duracion_host'] or '-').border = thin_border
    ws3.cell(row=row, column=7).alignment = center
    ws3.cell(row=row, column=8, value=clase['zoom_meetingid'] or '').border = thin_border
    
    estado_cell = ws3.cell(row=row, column=9, value=estados.get(clase['zoom_clasecompletada'], ''))
    estado_cell.border = thin_border
    estado_cell.alignment = center
    if clase['zoom_clasecompletada'] == 1:
        estado_cell.fill = green_fill
    elif clase['zoom_clasecompletada'] == 2:
        estado_cell.fill = red_fill
    
    row += 1

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 8
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 28
ws3.column_dimensions['E'].width = 22
ws3.column_dimensions['F'].width = 25
ws3.column_dimensions['G'].width = 12
ws3.column_dimensions['H'].width = 14
ws3.column_dimensions['I'].width = 12

ws3.freeze_panes = 'A2'

# Guardar
filename = f"Reporte_Profesores_CESCE_{fecha_inicio}_a_{fecha_fin}.xlsx"
filepath = f"/home/aulatuspeaking/www/app/moodle/reportes_cesce/{filename}"
wb.save(filepath)

print(f"\n✅ Reporte generado: {filepath}")
print(f"   Total clases: {len(clases)}")
print(f"   Clases con profesor: {clases_con_profesor} ({round(100*clases_con_profesor/len(clases))}%)")
print(f"   Clases sin profesor: {len(clases_sin_profesor)}")
print(f"   Profesores únicos: {len(profesores_stats)}")

cursor.close()
conn.close()
