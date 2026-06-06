#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Reporte FUNDAE CESCE - Compliance
"""

import sys
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

DB_CONFIG = {
    'host': 'localhost',
    'database': 'aulatuspeaking35',
    'user': 'moodle35',
    'password': 'TuspeakingFix2025!',
    'charset': 'utf8mb4'
}

def get_connection():
    return mysql.connector.connect(**DB_CONFIG)

def get_fundae_data(fecha_inicio, fecha_fin):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    
    query = """
    SELECT 
        'CESCE' as empresa,
        ca.acuity_type as tema,
        'Reunión' as tipo,
        ca.zoom_meetingid as meeting_id,
        ca.acuity_calendar as nombre_anfitrion,
        COALESCE(
            (SELECT p2.participant_email 
             FROM mdl_coding_zoom_participants p2 
             WHERE p2.zoom_meetingid = ca.zoom_meetingid 
               AND p2.participant_email LIKE '%%tuspeaking%%'
               AND p2.duration_minutes >= 30
             ORDER BY p2.duration_minutes DESC
             LIMIT 1),
            ''
        ) as email_anfitrion,
        COALESCE(ca.zoom_starttime, (SELECT MIN(px.join_time) FROM mdl_coding_zoom_participants px WHERE px.zoom_meetingid = ca.zoom_meetingid)) as hora_inicio,
        COALESCE(ca.zoom_endtime, (SELECT MAX(px.leave_time) FROM mdl_coding_zoom_participants px WHERE px.zoom_meetingid = ca.zoom_meetingid)) as hora_finalizacion,
        (SELECT COUNT(DISTINCT p3.id) 
         FROM mdl_coding_zoom_participants p3 
         WHERE p3.zoom_meetingid = ca.zoom_meetingid) as participantes,
        ca.zoom_duration as duracion_minutos,
        (SELECT COALESCE(SUM(p4.duration_minutes), 0)
         FROM mdl_coding_zoom_participants p4 
         WHERE p4.zoom_meetingid = ca.zoom_meetingid) as total_minutos_participantes,
        DATE(ca.acuity_datetime) as fecha_creacion,
        p.participant_name as nombre_participante,
        COALESCE(p.participant_email, '') as email_participante,
        p.join_time as hora_entrada,
        p.leave_time as hora_salida,
        p.duration_minutes as duracion_participante_min,
        CASE WHEN p.is_guest = 'Sí' OR p.is_guest = 'Yes' OR p.is_guest = '1' THEN 'Sí' ELSE 'No' END as invitado,
        ca.acuity_firstname as nombre_alumno_reserva,
        ca.acuity_lastname as apellido_alumno_reserva,
        ca.acuity_type as curso
    FROM mdl_cesce_acuityZoom ca
    INNER JOIN mdl_coding_zoom_participants p ON p.zoom_meetingid = ca.zoom_meetingid
    WHERE ca.acuity_datetime >= %s
      AND ca.acuity_datetime <= %s
      AND ca.acuity_firstname != 'Booking'
      AND ca.zoom_meetingid IS NOT NULL
    ORDER BY ca.acuity_datetime, ca.zoom_meetingid, p.join_time
    """
    
    cursor.execute(query, (fecha_inicio, fecha_fin))
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return data

def get_meetings_without_zoom_data(fecha_inicio, fecha_fin):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    
    query = """
    SELECT DISTINCT
        'CESCE' as empresa,
        ca.acuity_type as tema,
        'Reunión' as tipo,
        ca.zoom_meetingid as meeting_id,
        ca.acuity_calendar as nombre_anfitrion,
        '' as email_anfitrion,
        COALESCE(ca.zoom_starttime, (SELECT MIN(px.join_time) FROM mdl_coding_zoom_participants px WHERE px.zoom_meetingid = ca.zoom_meetingid)) as hora_inicio,
        COALESCE(ca.zoom_endtime, (SELECT MAX(px.leave_time) FROM mdl_coding_zoom_participants px WHERE px.zoom_meetingid = ca.zoom_meetingid)) as hora_finalizacion,
        0 as participantes,
        ca.zoom_duration as duracion_minutos,
        0 as total_minutos_participantes,
        DATE(ca.acuity_datetime) as fecha_creacion,
        '(Sin datos Zoom)' as nombre_participante,
        '' as email_participante,
        NULL as hora_entrada,
        NULL as hora_salida,
        0 as duracion_participante_min,
        '' as invitado,
        ca.acuity_firstname as nombre_alumno_reserva,
        ca.acuity_lastname as apellido_alumno_reserva,
        ca.acuity_type as curso
    FROM mdl_cesce_acuityZoom ca
    LEFT JOIN mdl_coding_zoom_participants p ON p.zoom_meetingid = ca.zoom_meetingid
    WHERE ca.acuity_datetime >= %s
      AND ca.acuity_datetime <= %s
      AND ca.acuity_firstname != 'Booking'
      AND ca.zoom_meetingid IS NOT NULL
      AND p.id IS NULL
    GROUP BY ca.zoom_meetingid, ca.acuity_email
    ORDER BY ca.acuity_datetime
    """
    
    cursor.execute(query, (fecha_inicio, fecha_fin))
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return data

def format_datetime(dt):
    if dt is None:
        return ''
    if isinstance(dt, str):
        if dt == '' or dt == 'None':
            return ''
        try:
            for fmt in ['%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                try:
                    dt = datetime.strptime(dt.split('+')[0].split('.')[0], fmt)
                    break
                except:
                    continue
            else:
                return dt
        except:
            return dt
    return dt.strftime('%m/%d/%Y %I:%M:%S %p')

def create_fundae_report(fecha_inicio, fecha_fin, output_path):
    print("=" * 60)
    print("  REPORTE FUNDAE CESCE - COMPLIANCE")
    print(f"  Periodo: {fecha_inicio} a {fecha_fin}")
    print("=" * 60)
    
    print("\nObteniendo datos...")
    data_con_zoom = get_fundae_data(fecha_inicio, fecha_fin)
    print(f"  Con participantes: {len(data_con_zoom)}")
    
    data_sin_zoom = get_meetings_without_zoom_data(fecha_inicio, fecha_fin)
    print(f"  Sin datos Zoom: {len(data_sin_zoom)}")
    
    all_data = list(data_con_zoom) + list(data_sin_zoom)
    print(f"Total: {len(all_data)}")
    
    if len(all_data) == 0:
        print("No hay datos")
        return 0
    
    wb = Workbook()
    ws = wb.active
    ws.title = "report"
    
    headers = ['empresa','tema','tipo','meeting_id','nombre_anfitrion','email_anfitrion',
               'hora_inicio','hora_finalizacion','participantes','duracion_minutos',
               'total_minutos_participantes','fecha_creacion','nombre_participante',
               'email_participante','hora_entrada','hora_salida','duracion_participante_min',
               'invitado','nombre_alumno_reserva','apellido_alumno_reserva','curso']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_idx = 2
    for record in all_data:
        ws.cell(row=row_idx, column=1, value=record.get('empresa', 'CESCE'))
        ws.cell(row=row_idx, column=2, value=record.get('tema', ''))
        ws.cell(row=row_idx, column=3, value=record.get('tipo', 'Reunión'))
        ws.cell(row=row_idx, column=4, value=record.get('meeting_id', ''))
        ws.cell(row=row_idx, column=5, value=record.get('nombre_anfitrion', ''))
        ws.cell(row=row_idx, column=6, value=record.get('email_anfitrion', ''))
        ws.cell(row=row_idx, column=7, value=format_datetime(record.get('hora_inicio')))
        ws.cell(row=row_idx, column=8, value=format_datetime(record.get('hora_finalizacion')))
        ws.cell(row=row_idx, column=9, value=record.get('participantes', 0))
        ws.cell(row=row_idx, column=10, value=record.get('duracion_minutos', 0) or 0)
        ws.cell(row=row_idx, column=11, value=record.get('total_minutos_participantes', 0) or 0)
        ws.cell(row=row_idx, column=12, value=format_datetime(record.get('fecha_creacion')))
        ws.cell(row=row_idx, column=13, value=record.get('nombre_participante', ''))
        ws.cell(row=row_idx, column=14, value=record.get('email_participante', ''))
        ws.cell(row=row_idx, column=15, value=format_datetime(record.get('hora_entrada')))
        ws.cell(row=row_idx, column=16, value=format_datetime(record.get('hora_salida')))
        ws.cell(row=row_idx, column=17, value=record.get('duracion_participante_min', 0) or 0)
        ws.cell(row=row_idx, column=18, value=record.get('invitado', ''))
        ws.cell(row=row_idx, column=19, value=record.get('nombre_alumno_reserva', ''))
        ws.cell(row=row_idx, column=20, value=record.get('apellido_alumno_reserva', ''))
        ws.cell(row=row_idx, column=21, value=record.get('curso', ''))
        row_idx += 1
    
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:U{row_idx - 1}"
    wb.save(output_path)
    
    print(f"\nReporte: {output_path}")
    print(f"Registros: {row_idx - 2}")
    return row_idx - 2

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Uso: python3 reporte_fundae_cesce.py FECHA_INICIO FECHA_FIN")
        sys.exit(1)
    
    fecha_inicio = sys.argv[1]
    fecha_fin = sys.argv[2]
    output_dir = "/home/aulatuspeaking/www/app/moodle/reportes_cesce"
    output_file = f"{output_dir}/Reporte_FUNDAE_CESCE_{fecha_inicio}_a_{fecha_fin}.xlsx"
    create_fundae_report(fecha_inicio, fecha_fin, output_file)
